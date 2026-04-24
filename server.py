import csv
import os
import io
import json
import base64
import hashlib
import secrets
import functools
import re
import tempfile
import smtplib
import threading
from datetime import datetime, date, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage

from flask import Flask, request, jsonify, send_from_directory, send_file
import qrcode
import pyotp
from io import BytesIO

# ============================================================
# CONFIG
# ============================================================

app = Flask(__name__, static_folder='public', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_POSTGRES = DATABASE_URL.startswith('postgres')
ADMIN_DEFAULT_PIN = os.environ.get('ADMIN_PIN', '1234')
IS_PRODUCTION = os.environ.get('RENDER', '') or os.environ.get('PRODUCTION', '')
SMTP_USER = os.environ.get('SMTP_USER', 'ptcpob@gmail.com')
SMTP_PASS = os.environ.get('SMTP_PASS', '')
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))

ROSTER_CSV = os.path.join(os.path.dirname(__file__), 'adnoc_workforce_roste.csv')

ROLE_HIERARCHY = {
    'executive': 50,
    'admin': 40,
    'manager': 35,
    'project_manager': 30,
    'focal_point': 25,
    'supervisor': 20,
    'scanner': 18,
    'viewer': 10
}

# Hierarchy: Division -> Area -> Project -> Personnel
# 4 divisions; each has areas; each area has projects; headcount per project
DIVISIONS = [
    ('P&C BAB/NEB', ['BAB', 'NEB', 'BAB MP']),
    ('P&C GTG', ['GAS', 'TPO', 'GAS/TPO', 'WEP', 'FUJ', 'GAS (ASAB)', 'GAS (BAB)', 'IPS', 'JD', 'MPS']),
    ('P&C SE', ['Asab/Sahil', 'SQM', 'SHAH', 'QW', 'MN']),
    ('P&C BUHASA', ['Buhasa', 'BUIFDP (Buhasa)', 'BUHASA MP']),
]

AREA_ALIASES = {
    'P&C SHAH': 'SHAH', 'P&C (SHAH)': 'SHAH', 'Shah': 'SHAH', 'shah': 'SHAH',
    'P&C Qusahwira': 'QW', 'P&C (Qusahwira/Mender QW/MN)': 'QW', 'Qusahwira': 'QW',
    'P&C Mender': 'MN', 'Mender': 'MN',
    'P&C (FUJ)': 'FUJ', 'P&C FUJ': 'FUJ', 'fuj': 'FUJ', 'Fuj': 'FUJ',
    'P&C (IPS)': 'IPS', '(IPS)': 'IPS',
    'P&C (JD)': 'JD', '(JD)': 'JD',
    'P&C (MPS)': 'MPS', '(MPS)': 'MPS',
    'P&C GAS(BAB)': 'GAS (BAB)', 'GAS(BAB)': 'GAS (BAB)',
    'P&C GAS(ASAB)': 'GAS (ASAB)', 'GAS(ASAB)': 'GAS (ASAB)',
    'P&C Asab': 'GAS (ASAB)',
    'P&C Sahil': 'SHAH',
    'P&C Buhasa': 'BUHASA', 'Buhasa': 'BUHASA', 'buhasa': 'BUHASA',
    'Buhasa MP': 'BUHASA MP', 'P&C Buhasa MP': 'BUHASA MP',
    'BAB MP': 'BAB MP',
}

COMMON_PINS = {'0000', '1111', '2222', '3333', '4444', '5555', '6666', '7777', '8888', '9999',
               '1234', '4321', '1122', '1212', '0123', '9876', '5678', '8765'}

LOGIN_ATTEMPTS = {}
LOGIN_WINDOW = 60
LOGIN_MAX_PER_IP = 20

# ============================================================
# HTTPS REDIRECT (when behind proxy e.g. Render)
# ============================================================

@app.before_request
def redirect_http_to_https():
    """Redirect HTTP to HTTPS when the app is behind a proxy that sets X-Forwarded-Proto."""
    if not request.is_secure and request.headers.get('X-Forwarded-Proto') == 'http':
        from flask import redirect
        return redirect(request.url.replace('http://', 'https://', 1), code=301)


# ============================================================
# SECURITY HEADERS
# ============================================================

@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' https://unpkg.com; "
        "style-src 'self' https://fonts.googleapis.com 'unsafe-inline'; "
        "font-src https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'"
    )
    return response

# ============================================================
# DATABASE ABSTRACTION (SQLite local / PostgreSQL cloud)
# ============================================================

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras

    def get_db():
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = False
        return conn

    def _translate_sql(sql):
        sql = sql.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'SERIAL PRIMARY KEY')
        sql = sql.replace('TIMESTAMP DEFAULT CURRENT_TIMESTAMP', "TIMESTAMP DEFAULT NOW()")
        # Order matters: replace longer patterns first
        sql = sql.replace("datetime('now', '+24 hours')", "NOW() + INTERVAL '24 hours'")
        sql = sql.replace("datetime('now', '+5 minutes')", "NOW() + INTERVAL '5 minutes'")
        sql = sql.replace(
            "datetime('now', '+' || ? || ' minutes')",
            "(NOW() + (%s || ' minutes')::interval)"
        )
        sql = sql.replace("datetime('now')", "NOW()")
        sql = sql.replace('?', '%s')
        sql = re.sub(r'INSERT OR REPLACE INTO', 'INSERT INTO', sql)
        if 'INSERT OR IGNORE INTO' in sql:
            sql = sql.replace('INSERT OR IGNORE INTO', 'INSERT INTO')
            if 'ON CONFLICT' not in sql:
                sql = re.sub(r'(VALUES\s*\([^)]+\))', r'\1 ON CONFLICT DO NOTHING', sql)
        return sql

    def db_execute(conn, sql, params=None):
        sql = _translate_sql(sql)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params or [])
        return cur

    def db_fetchone(conn, sql, params=None):
        cur = db_execute(conn, sql, params)
        return cur.fetchone()

    def db_fetchall(conn, sql, params=None):
        cur = db_execute(conn, sql, params)
        return cur.fetchall()

else:
    import sqlite3

    def get_db():
        conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), 'pob.db'))
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        return conn

    def db_execute(conn, sql, params=None):
        return conn.execute(sql, params or [])

    def db_fetchone(conn, sql, params=None):
        row = conn.execute(sql, params or []).fetchone()
        return dict(row) if row else None

    def db_fetchall(conn, sql, params=None):
        return [dict(r) for r in conn.execute(sql, params or []).fetchall()]


def hash_pin(pin):
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', pin.encode(), salt.encode(), 100000)
    return f"{salt}:{h.hex()}"


def verify_pin(pin, stored):
    if ':' not in stored:
        return hashlib.sha256(pin.encode()).hexdigest() == stored
    salt, h = stored.split(':', 1)
    return hashlib.pbkdf2_hmac('sha256', pin.encode(), salt.encode(), 100000).hex() == h


def is_weak_pin(pin):
    if pin in COMMON_PINS:
        return True
    if len(set(pin)) == 1:
        return True
    digits = [int(d) for d in pin]
    if all(digits[i+1] - digits[i] == 1 for i in range(len(digits)-1)):
        return True
    if all(digits[i] - digits[i+1] == 1 for i in range(len(digits)-1)):
        return True
    return False


# ============================================================
# DATABASE INIT
# ============================================================

def init_db():
    conn = get_db()
    try:
        tables = [
            '''CREATE TABLE IF NOT EXISTS divisions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS areas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                division_id INTEGER NOT NULL REFERENCES divisions(id),
                name TEXT NOT NULL,
                active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(division_id, name)
            )''',
            '''CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                area_id INTEGER REFERENCES areas(id),
                name TEXT NOT NULL,
                description TEXT,
                agreement_no TEXT,
                contract_number TEXT,
                contractor_company TEXT,
                active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(area_id, name)
            )''',
            '''CREATE TABLE IF NOT EXISTS sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                project_id INTEGER NOT NULL REFERENCES projects(id),
                description TEXT,
                active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(name, project_id)
            )''',
            '''CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                srl TEXT, agreement_no TEXT, name TEXT NOT NULL, nationality TEXT,
                dob TEXT, designation TEXT, work_location TEXT, camp_name TEXT,
                employee_no TEXT NOT NULL, qualification TEXT, date_joining TEXT,
                date_deployment TEXT, medical_date TEXT, discipline TEXT,
                subcontractor TEXT, remarks TEXT,
                asset_name TEXT, contractor TEXT, age TEXT,
                eid_passport TEXT, fieldglass_status TEXT,
                medical_frequency TEXT, last_medical_date TEXT,
                next_medical_due TEXT, medical_result TEXT,
                chronic_condition TEXT, chronic_treated TEXT,
                general_feeling TEXT,
                project_id INTEGER NOT NULL REFERENCES projects(id),
                active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(employee_no, project_id)
            )''',
            '''CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                display_name TEXT NOT NULL,
                pin_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'supervisor',
                active INTEGER DEFAULT 1,
                failed_attempts INTEGER DEFAULT 0,
                locked_until TIMESTAMP,
                must_change_pin INTEGER DEFAULT 0,
                totp_secret TEXT,
                totp_enabled INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS user_project_access (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                project_id INTEGER NOT NULL REFERENCES projects(id),
                UNIQUE(user_id, project_id)
            )''',
            '''CREATE TABLE IF NOT EXISTS user_division_access (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                division_id INTEGER NOT NULL REFERENCES divisions(id),
                UNIQUE(user_id, division_id)
            )''',
            '''CREATE TABLE IF NOT EXISTS auth_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                token TEXT UNIQUE NOT NULL,
                device_info TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL
            )''',
            '''CREATE TABLE IF NOT EXISTS pending_2fa (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                pending_token TEXT UNIQUE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL
            )''',
            '''CREATE TABLE IF NOT EXISTS download_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                token TEXT UNIQUE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL
            )''',
            '''CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT NOT NULL,
                detail TEXT,
                ip_address TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                employee_no TEXT NOT NULL,
                project_id INTEGER NOT NULL REFERENCES projects(id),
                site_id INTEGER NOT NULL REFERENCES sites(id),
                scan_date TEXT NOT NULL,
                session TEXT NOT NULL,
                scanned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                supervisor_id INTEGER REFERENCES users(id),
                supervisor_name TEXT,
                latitude REAL, longitude REAL,
                UNIQUE(employee_no, project_id, scan_date, session)
            )''',
            '''CREATE TABLE IF NOT EXISTS twl_readings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                area_id INTEGER REFERENCES areas(id),
                site_id INTEGER REFERENCES sites(id),
                reading_date TEXT NOT NULL,
                reading_time TEXT,
                twl_value REAL NOT NULL,
                risk_zone TEXT NOT NULL,
                temperature REAL,
                humidity REAL,
                wind_speed REAL,
                work_type TEXT DEFAULT 'light',
                recorded_by INTEGER REFERENCES users(id),
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''',
            '''CREATE TABLE IF NOT EXISTS observations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                division_id INTEGER REFERENCES divisions(id),
                area_id INTEGER REFERENCES areas(id),
                project_id INTEGER REFERENCES projects(id),
                observation_date TEXT NOT NULL,
                observer_name TEXT,
                observer_designation TEXT,
                observer_discipline TEXT,
                observer_company TEXT,
                employee_type TEXT,
                observation_group TEXT NOT NULL,
                observation_type TEXT,
                potential_severity TEXT,
                risk_rating TEXT,
                observation_text TEXT,
                corrective_action TEXT,
                intervention TEXT,
                outcome TEXT,
                remarks TEXT,
                recorded_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )'''
        ]

        for sql in tables:
            try:
                db_execute(conn, sql)
                conn.commit()
            except Exception as e:
                conn.rollback()
                print(f"  Table warning: {e}")

        indexes = [
            'CREATE INDEX IF NOT EXISTS idx_att_date ON attendance(scan_date)',
            'CREATE INDEX IF NOT EXISTS idx_att_project ON attendance(project_id)',
            'CREATE INDEX IF NOT EXISTS idx_att_site ON attendance(site_id)',
            'CREATE INDEX IF NOT EXISTS idx_emp_no ON employees(employee_no)',
            'CREATE INDEX IF NOT EXISTS idx_emp_project ON employees(project_id)',
            'CREATE INDEX IF NOT EXISTS idx_projects_area ON projects(area_id)',
            'CREATE INDEX IF NOT EXISTS idx_areas_division ON areas(division_id)',
            'CREATE INDEX IF NOT EXISTS idx_token ON auth_tokens(token)',
            'CREATE INDEX IF NOT EXISTS idx_upa_user ON user_project_access(user_id)',
            'CREATE INDEX IF NOT EXISTS idx_uda_user ON user_division_access(user_id)',
            'CREATE INDEX IF NOT EXISTS idx_twl_date ON twl_readings(reading_date)',
            'CREATE INDEX IF NOT EXISTS idx_twl_area ON twl_readings(area_id)',
            'CREATE INDEX IF NOT EXISTS idx_obs_date ON observations(observation_date)',
            'CREATE INDEX IF NOT EXISTS idx_obs_division ON observations(division_id)',
            'CREATE INDEX IF NOT EXISTS idx_obs_group ON observations(observation_group)',
        ]
        for sql in indexes:
            try:
                db_execute(conn, sql)
                conn.commit()
            except Exception:
                conn.rollback()

        # Add new columns if missing (migrations)
        for alter in [
            'ALTER TABLE projects ADD COLUMN region TEXT',
            'ALTER TABLE projects ADD COLUMN area_id INTEGER REFERENCES areas(id)',
            'ALTER TABLE projects ADD COLUMN contract_number TEXT',
            'ALTER TABLE projects ADD COLUMN contractor_company TEXT',
            'ALTER TABLE users ADD COLUMN email TEXT',
            'ALTER TABLE users ADD COLUMN designation TEXT',
            'ALTER TABLE attendance ADD COLUMN scanner_email TEXT',
            'ALTER TABLE attendance ADD COLUMN scanner_designation TEXT',
            'ALTER TABLE employees ADD COLUMN asset_name TEXT',
            'ALTER TABLE employees ADD COLUMN contractor TEXT',
            'ALTER TABLE employees ADD COLUMN age TEXT',
            'ALTER TABLE employees ADD COLUMN eid_passport TEXT',
            'ALTER TABLE employees ADD COLUMN fieldglass_status TEXT',
            'ALTER TABLE employees ADD COLUMN medical_frequency TEXT',
            'ALTER TABLE employees ADD COLUMN last_medical_date TEXT',
            'ALTER TABLE employees ADD COLUMN next_medical_due TEXT',
            'ALTER TABLE employees ADD COLUMN medical_result TEXT',
            'ALTER TABLE employees ADD COLUMN chronic_condition TEXT',
            'ALTER TABLE employees ADD COLUMN chronic_treated TEXT',
            'ALTER TABLE employees ADD COLUMN general_feeling TEXT',
        ]:
            try:
                db_execute(conn, alter)
                conn.commit()
            except Exception:
                conn.rollback()

        # Seed divisions and areas (Division -> Area -> Project)
        for div_name, area_names in DIVISIONS:
            row = db_fetchone(conn, 'SELECT id FROM divisions WHERE name = ?', (div_name,))
            if not row:
                db_execute(conn, 'INSERT INTO divisions (name, active) VALUES (?, 1)', (div_name,))
                conn.commit()
                row = db_fetchone(conn, 'SELECT id FROM divisions WHERE name = ?', (div_name,))
            div_id = row['id'] if isinstance(row, dict) else row[0]
            for area_name in area_names:
                ar = db_fetchone(conn, 'SELECT id FROM areas WHERE division_id = ? AND name = ?', (div_id, area_name))
                if not ar:
                    db_execute(conn, 'INSERT INTO areas (division_id, name, active) VALUES (?, ?, 1)', (div_id, area_name))
                    conn.commit()
                    ar = db_fetchone(conn, 'SELECT id FROM areas WHERE division_id = ? AND name = ?', (div_id, area_name))
                area_id = ar['id'] if isinstance(ar, dict) else ar[0]
        # Normalize old area names to canonical names
        for old_name, canonical in AREA_ALIASES.items():
            try:
                existing = db_fetchone(conn, 'SELECT id FROM areas WHERE name = ?', (old_name,))
                if existing:
                    canon_exists = db_fetchone(conn, 'SELECT id FROM areas WHERE name = ? AND id != ?',
                                               (canonical, existing['id'] if isinstance(existing, dict) else existing[0]))
                    if canon_exists:
                        old_id = existing['id'] if isinstance(existing, dict) else existing[0]
                        canon_id = canon_exists['id'] if isinstance(canon_exists, dict) else canon_exists[0]
                        db_execute(conn, 'UPDATE projects SET area_id = ? WHERE area_id = ?', (canon_id, old_id))
                        db_execute(conn, 'UPDATE user_division_access SET division_id = division_id WHERE 1=0')
                        db_execute(conn, 'DELETE FROM areas WHERE id = ?', (old_id,))
                    else:
                        db_execute(conn, 'UPDATE areas SET name = ? WHERE id = ?',
                                   (canonical, existing['id'] if isinstance(existing, dict) else existing[0]))
                    conn.commit()
            except Exception:
                conn.rollback()

        # Backfill area_id for existing projects that have region set (legacy)
        try:
            for div_name, area_names in DIVISIONS:
                div_row = db_fetchone(conn, 'SELECT id FROM divisions WHERE name = ?', (div_name,))
                if not div_row:
                    continue
                div_id = div_row['id'] if isinstance(div_row, dict) else div_row[0]
                for area_name in area_names:
                    ar = db_fetchone(conn, 'SELECT id FROM areas WHERE division_id = ? AND name = ?', (div_id, area_name))
                    if not ar:
                        continue
                    area_id = ar['id'] if isinstance(ar, dict) else ar[0]
                    db_execute(conn, 'UPDATE projects SET area_id = ? WHERE (area_id IS NULL OR area_id = 0) AND (name = ? OR region = ?)', (area_id, area_name, div_name))
                    conn.commit()
        except Exception:
            conn.rollback()

        # Seed projects from contractors CSV (if available)
        csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'contractors.csv')
        if os.path.exists(csv_path):
            import csv
            try:
                with open(csv_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        proj_name = (row.get('Project Name') or '').strip()
                        contract_no = (row.get('Contract Number') or '').strip()
                        div_name = (row.get('Division') or '').strip()
                        area_name = (row.get('Area') or '').strip()
                        contractor = (row.get('Name of Contractors') or '').strip()
                        if not proj_name or not div_name or not area_name:
                            continue
                        div_row = db_fetchone(conn, 'SELECT id FROM divisions WHERE name = ?', (div_name,))
                        if not div_row:
                            continue
                        div_id = div_row['id'] if isinstance(div_row, dict) else div_row[0]
                        ar = db_fetchone(conn, 'SELECT id FROM areas WHERE division_id = ? AND name = ?', (div_id, area_name))
                        if not ar:
                            continue
                        area_id = ar['id'] if isinstance(ar, dict) else ar[0]
                        existing = db_fetchone(conn, 'SELECT id FROM projects WHERE area_id = ? AND name = ?', (area_id, proj_name))
                        if existing:
                            pid = existing['id'] if isinstance(existing, dict) else existing[0]
                            db_execute(conn, 'UPDATE projects SET contract_number = ?, contractor_company = ? WHERE id = ?',
                                       (contract_no, contractor, pid))
                        else:
                            try:
                                db_execute(conn, '''INSERT INTO projects (area_id, name, contract_number, contractor_company, active)
                                    VALUES (?, ?, ?, ?, 1)''', (area_id, proj_name, contract_no, contractor))
                            except Exception:
                                conn.rollback()
                                continue
                        conn.commit()
                print(f'  Seeded projects from contractors CSV')
            except Exception as e:
                print(f'  CSV seed warning: {e}')
                conn.rollback()

        user = db_fetchone(conn, 'SELECT COUNT(*) as c FROM users')
        count = user['c'] if isinstance(user, dict) else user[0]
        if count == 0:
            totp_secret = pyotp.random_base32()
            db_execute(conn, '''
                INSERT INTO users (username, display_name, pin_hash, role, must_change_pin, totp_secret, totp_enabled)
                VALUES (?, ?, ?, 'executive', 0, ?, 1)
            ''', ('admin', 'Administrator', hash_pin('0000'), totp_secret))
            conn.commit()
            uri = pyotp.TOTP(totp_secret).provisioning_uri(name='admin', issuer_name='PTC POB Tracker')
            print('  Default admin (2FA only): username=admin')
            print(f'  Add this secret to your authenticator app: {totp_secret}')
            print(f'  Or scan QR for: {uri[:60]}...')
        else:
            # Non-scanner users without 2FA get a secret so admin login works
            rows = db_fetchall(conn, "SELECT id, username, role FROM users WHERE (totp_secret IS NULL OR totp_enabled = 0) AND role != 'scanner'")
            for row in rows:
                uid = row['id'] if isinstance(row, dict) else row[0]
                uname = (row.get('username') or row[1]) if isinstance(row, dict) else row[1]
                new_secret = pyotp.random_base32()
                db_execute(conn, 'UPDATE users SET totp_secret = ?, totp_enabled = 1 WHERE id = ?', (new_secret, uid))
                conn.commit()
                print(f'  2FA set for user {uname}. Add to authenticator: {new_secret}')

        # ── Anonymize worker data (one-time migration for demo/POC) ──
        _anonymize_workers(conn)

        # ── Seed demo attendance + O&I if tables are empty ──
        _seed_demo_attendance(conn)
        _seed_demo_observations(conn)
        _seed_demo_twl(conn)

    finally:
        conn.close()


def _anonymize_workers(conn):
    """Replace ALL employee personal data with dummy data. Runs once, skips if already done."""
    import random as _rnd
    from datetime import datetime as _dt, timedelta as _td
    _rnd.seed(42)

    try:
        marker = db_fetchone(conn, "SELECT COUNT(*) as c FROM employees WHERE employee_no LIKE ?", ('DEM-%',))
        already = marker['c'] if isinstance(marker, dict) else marker[0]
        if already > 10:
            print(f'  Anonymization already done ({already} DEM- records), skipping.')
            return
    except Exception as e:
        print(f'  Anonymization marker check failed: {e}')
        try:
            conn.rollback()
        except Exception:
            pass
        return

    total_row = db_fetchone(conn, "SELECT COUNT(*) as c FROM employees")
    total = total_row['c'] if isinstance(total_row, dict) else total_row[0]
    if total == 0:
        return

    print(f'  Anonymizing {total} employees (all rows)...')

    FIRST = {
        'Indian': ['Rajesh','Suresh','Anil','Vikram','Sanjay','Mahesh','Deepak','Ravi','Amit','Ajay',
                    'Pradeep','Manoj','Ramesh','Venkat','Ganesh','Ashok','Sunil','Naresh','Kamal','Mukesh',
                    'Dinesh','Kishore','Mohan','Gopal','Harish','Satish','Vinod','Jitendra','Bhaskar','Naveen'],
        'Pakistani': ['Ali','Hassan','Usman','Bilal','Imran','Tariq','Faisal','Nadeem','Shahid','Kamran',
                      'Waseem','Rizwan','Zafar','Irfan','Asif','Saleem','Arif','Khalid','Javed','Nasir'],
        'Filipino': ['Jose','Juan','Mark','Joel','Ryan','Carlo','Angelo','Rodel','Noel','Dennis',
                     'Ronald','Michael','Jason','Allan','Leo','Arnel','Gilbert','Rommel','Jayson','Elmer'],
        'Bangladeshi': ['Mohammad','Abdul','Karim','Rahim','Hasan','Jamal','Fazlul','Mizanur','Shahidul','Rafiq',
                        'Monir','Sohel','Liton','Rubel','Habib','Alamgir','Shafiq','Mostafa','Rashed','Zahid'],
        'Egyptian': ['Ahmed','Mohamed','Mahmoud','Mostafa','Ibrahim','Khaled','Youssef','Omar','Tarek','Hesham',
                     'Waleed','Essam','Sherif','Hany','Adel','Samir','Nabil','Gamal','Ashraf','Magdy'],
        'Nepalese': ['Ram','Shyam','Hari','Bishnu','Krishna','Gopal','Dipak','Suman','Bikram','Prakash',
                     'Binod','Santosh','Rajendra','Narayan','Bhim','Tek','Purna','Dhan','Keshav','Ganga'],
    }
    LAST = {
        'Indian': ['Kumar','Singh','Sharma','Patel','Rao','Reddy','Nair','Pillai','Iyer','Gupta',
                    'Verma','Mishra','Das','Jha','Yadav','Chauhan','Thakur','Pandey','Bose','Sinha'],
        'Pakistani': ['Khan','Ahmed','Hussain','Malik','Iqbal','Butt','Chaudhry','Sheikh','Qureshi','Siddiqui',
                      'Raza','Abbasi','Mirza','Bhatti','Aslam','Rehman','Naeem','Baig','Akram','Anwar'],
        'Filipino': ['Santos','Reyes','Cruz','Garcia','Lopez','Torres','Ramos','Flores','Rivera','Gonzales',
                     'Bautista','Aquino','Mendoza','Castillo','Villanueva','Dela Cruz','Navarro','Mercado','Pascual','Soriano'],
        'Bangladeshi': ['Hossain','Islam','Rahman','Akter','Mia','Khatun','Uddin','Begum','Sarker','Ali',
                        'Chowdhury','Siddique','Talukdar','Biswas','Bhuiyan','Kabir','Haque','Amin','Sultana','Khan'],
        'Egyptian': ['El-Sayed','Hassan','Ali','Abdel-Fattah','Farouk','Mansour','Salah','Nasser','Osman','Ismail',
                     'Darwish','Helmy','Saad','Attia','Ramadan','Soliman','Fouad','Moussa','Rizk','Shawky'],
        'Nepalese': ['Gurung','Tamang','Rai','Magar','Thapa','Shrestha','Lama','Adhikari','Karki','Limbu',
                     'Chhetri','Poudel','Bhandari','Khadka','Basnet','Bhattarai','Sapkota','Regmi','Subedi','Dahal'],
    }
    NATS = ['Indian','Pakistani','Filipino','Bangladeshi','Egyptian','Nepalese']
    NAT_W = [35,20,15,15,10,5]
    DESIGS = ['Welder','Pipefitter','Electrician','Rigger','Scaffolder','Mechanic','Painter','Insulator',
              'Carpenter','Mason','Foreman','Technician','Safety Officer','Crane Operator',
              'Instrument Technician','Supervisor','QC Inspector','Store Keeper','Driver','Helper','Fitter','Fabricator']
    DISCS = ['Civil','Mechanical','Electrical','Piping','Instrumentation','HSE','Structural','Welding',
             'Painting','Insulation','Scaffolding','QA/QC','Rigging','Operations','Logistics']
    QUALS = ['ITI','Diploma','B.Tech','High School','Trade Certificate','NCVT','BSc','Intermediate',
             '10th Pass','Graduate','Certification','NVQ Level 3','Technical Diploma','CSWIP']
    CAMPS = ['ICAD Worker Village','Musaffah Camp','Ruwais Housing Complex','Bu Hasa Camp','Habshan Camp',
             'Asab Camp','Shah Gas Camp','Jebel Dhanna Camp','Bab Field Camp','NEB Worker Camp',
             'Gayathi Camp','Tarif Camp','Madinat Zayed Camp']
    MED_RESULTS = ['Fit','Fit','Fit','Fit','Fit','Fit','Fit','UNFIT','Pending','']
    MED_FREQ = ['1 Year','2 Years','1 Year','1 Year','6 Months','']
    CHRONIC = ['','','','','','','','Diabetes','Hypertension','Cholesterol','Asthma','Back Pain (Chronic)','Nil','None']
    CHRONIC_TR = ['','','Yes','Yes - Under Control','No','N/A']
    FEELINGS = ['Good','Good','Good','Good','Fine','Fine','Excellent','','Tired','OK']
    FIELDGLASS = ['Active','Active','Active','Pending','Inactive','']
    ASSETS = ['Asset-A','Asset-B','Asset-C','Asset-D','']
    REMARKS = ['','','','','','New joiner','Transferred from other site','Experienced worker','']

    def _rand_date(sy=1975, ey=2000):
        s = _dt(sy,1,1); d = (_dt(ey,12,31)-s).days
        return (s + _td(days=_rnd.randint(0,d))).strftime('%Y-%m-%d')
    def _calc_age(dob_str):
        try:
            dob = _dt.strptime(dob_str, '%Y-%m-%d')
            today = _dt.now()
            return str(today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day)))
        except Exception:
            return ''
    def _rand_recent(m=24):
        return (_dt.now() - _td(days=_rnd.randint(0,m*30))).strftime('%Y-%m-%d')
    def _rand_future(m=18):
        return (_dt.now() + _td(days=_rnd.randint(-60,m*30))).strftime('%Y-%m-%d')

    # First pass: set all employee_no to temporary unique values to avoid UNIQUE constraint conflicts
    try:
        db_execute(conn, "UPDATE employees SET employee_no = 'TMP-' || CAST(id AS TEXT)")
        conn.commit()
        print('  Cleared employee_no for re-assignment')
    except Exception as ex:
        print(f'  Clear employee_no failed: {ex}')
        try:
            conn.rollback()
        except Exception:
            pass
        return

    rows = db_fetchall(conn, 'SELECT id FROM employees')
    used = set()
    count = 0
    for row in rows:
        eid = row['id'] if isinstance(row, dict) else row[0]
        nat = _rnd.choices(NATS, weights=NAT_W, k=1)[0]
        name = f"{_rnd.choice(FIRST[nat])} {_rnd.choice(LAST[nat])}"
        emp_no = f"DEM-{_rnd.randint(10000,99999)}"
        while emp_no in used:
            emp_no = f"DEM-{_rnd.randint(10000,99999)}"
        used.add(emp_no)
        dob = _rand_date(1975, 2000)
        age = _calc_age(dob)
        chronic = _rnd.choice(CHRONIC)
        ct = _rnd.choice(CHRONIC_TR) if chronic and chronic not in ('','Nil','None') else ''
        agreement = f"AGR-{_rnd.randint(10000,99999)}"
        srl = str(_rnd.randint(1, 9999))

        try:
            db_execute(conn, '''UPDATE employees SET
                name=?, nationality=?, dob=?, age=?, designation=?, discipline=?,
                qualification=?, camp_name=?, employee_no=?, srl=?, agreement_no=?,
                date_joining=?, date_deployment=?, medical_date=?,
                last_medical_date=?, next_medical_due=?, medical_result=?,
                medical_frequency=?, chronic_condition=?, chronic_treated=?,
                general_feeling=?, eid_passport=?, fieldglass_status=?,
                asset_name=?, remarks=?
                WHERE id=?''',
                (name, nat, dob, age, _rnd.choice(DESIGS), _rnd.choice(DISCS),
                 _rnd.choice(QUALS), _rnd.choice(CAMPS), emp_no, srl, agreement,
                 _rand_recent(36), _rand_recent(24), _rand_recent(12),
                 _rand_recent(8), _rand_future(12), _rnd.choice(MED_RESULTS),
                 _rnd.choice(MED_FREQ), chronic, ct,
                 _rnd.choice(FEELINGS),
                 f"784-{_rnd.randint(1950,2005)}-{_rnd.randint(1000000,9999999)}-{_rnd.randint(1,9)}",
                 _rnd.choice(FIELDGLASS), _rnd.choice(ASSETS), _rnd.choice(REMARKS),
                 eid))
            count += 1
        except Exception as ex:
            print(f'  Anonymize row {eid} failed: {ex}')
            try:
                conn.rollback()
            except Exception:
                pass

    conn.commit()
    print(f'  Anonymized {count}/{len(rows)} employees')

    # Update attendance records to match new employee_nos
    try:
        att_rows = db_fetchall(conn, 'SELECT DISTINCT employee_id FROM attendance WHERE employee_id IS NOT NULL')
        att_count = 0
        for ar in att_rows:
            aid = ar['employee_id'] if isinstance(ar, dict) else ar[0]
            emp = db_fetchone(conn, 'SELECT employee_no FROM employees WHERE id = ?', (aid,))
            if emp:
                new_no = emp['employee_no'] if isinstance(emp, dict) else emp[0]
                db_execute(conn, 'UPDATE attendance SET employee_no = ? WHERE employee_id = ?', (new_no, aid))
                att_count += 1
        conn.commit()
        print(f'  Updated {att_count} attendance records')
    except Exception as ex:
        print(f'  Attendance update failed: {ex}')
        try:
            conn.rollback()
        except Exception:
            pass


def _seed_demo_attendance(conn):
    """Generate 60 days of realistic attendance data for all active employees."""
    import random as _rnd
    from datetime import date as _date, timedelta as _td
    _rnd.seed(99)

    existing = db_fetchone(conn, "SELECT COUNT(*) as c FROM attendance")
    cnt = existing['c'] if isinstance(existing, dict) else existing[0]
    if cnt > 100:
        print(f'  Attendance already seeded ({cnt} rows), skipping.')
        return

    employees = db_fetchall(conn, """
        SELECT e.id, e.employee_no, e.project_id, p.area_id
        FROM employees e JOIN projects p ON p.id = e.project_id
        WHERE e.active = 1
    """)
    if not employees:
        print('  No active employees for attendance seed.')
        return

    sites_map = {}
    sites = db_fetchall(conn, "SELECT id, project_id FROM sites")
    for s in sites:
        pid = s['project_id'] if isinstance(s, dict) else s[1]
        sid = s['id'] if isinstance(s, dict) else s[0]
        sites_map.setdefault(pid, []).append(sid)

    project_ids = set()
    for emp in employees:
        pid = emp['project_id'] if isinstance(emp, dict) else emp[2]
        project_ids.add(pid)
    for pid in project_ids:
        if pid not in sites_map:
            try:
                db_execute(conn, "INSERT INTO sites (name, project_id) VALUES (?, ?)", ('Main Site', pid))
            except Exception:
                pass
    conn.commit()
    sites = db_fetchall(conn, "SELECT id, project_id FROM sites")
    sites_map = {}
    for s in sites:
        pid = s['project_id'] if isinstance(s, dict) else s[1]
        sid = s['id'] if isinstance(s, dict) else s[0]
        sites_map.setdefault(pid, []).append(sid)

    today = _date.today()
    days = 60
    sessions = ['AM', 'EV']
    batch = []
    batch_size = 500

    for day_offset in range(days):
        d = today - _td(days=day_offset)
        if d.weekday() >= 5:
            continue
        ds = d.isoformat()

        for sess in sessions:
            am_rate = _rnd.uniform(0.70, 0.92)
            ev_rate = am_rate * _rnd.uniform(0.85, 0.98) if sess == 'EV' else am_rate

            rate = am_rate if sess == 'AM' else ev_rate
            for emp in employees:
                if _rnd.random() > rate:
                    continue
                eid = emp['id'] if isinstance(emp, dict) else emp[0]
                eno = emp['employee_no'] if isinstance(emp, dict) else emp[1]
                pid = emp['project_id'] if isinstance(emp, dict) else emp[2]
                site_list = sites_map.get(pid, list(sites_map.values())[0] if sites_map else [1])
                sid = _rnd.choice(site_list)

                batch.append((eid, eno, pid, sid, ds, sess))

                if len(batch) >= batch_size:
                    _flush_attendance_batch(conn, batch)
                    batch = []

    if batch:
        _flush_attendance_batch(conn, batch)

    final = db_fetchone(conn, "SELECT COUNT(*) as c FROM attendance")
    fc = final['c'] if isinstance(final, dict) else final[0]
    print(f'  Seeded {fc} attendance records over {days} weekdays')


def _flush_attendance_batch(conn, batch):
    for (eid, eno, pid, sid, ds, sess) in batch:
        try:
            db_execute(conn, """INSERT OR IGNORE INTO attendance
                (employee_id, employee_no, project_id, site_id, scan_date, session)
                VALUES (?, ?, ?, ?, ?, ?)""",
                (eid, eno, pid, sid, ds, sess))
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
    try:
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


def _seed_demo_observations(conn):
    """Generate 60 days of O&I observations linked to real project/area/division structure."""
    import random as _rnd
    from datetime import date as _date, timedelta as _td
    _rnd.seed(77)

    existing = db_fetchone(conn, "SELECT COUNT(*) as c FROM observations")
    cnt = existing['c'] if isinstance(existing, dict) else existing[0]
    if cnt > 20:
        print(f'  Observations already seeded ({cnt} rows), skipping.')
        return

    projects = db_fetchall(conn, """
        SELECT p.id, p.area_id, a.division_id, p.contractor_company
        FROM projects p
        JOIN areas a ON a.id = p.area_id
        WHERE p.active = 1
    """)
    if not projects:
        print('  No projects for observation seed.')
        return

    employees = db_fetchall(conn, """
        SELECT e.name, e.designation, e.discipline, p.contractor_company
        FROM employees e JOIN projects p ON p.id = e.project_id
        WHERE e.active = 1
    """)

    groups = ['Safe Act', 'Safe Condition', 'Unsafe Act', 'Unsafe Condition', 'Near Miss', 'HIPO']
    group_weights = [30, 25, 20, 15, 7, 3]
    types_list = [
        'Hot Work', 'Safety Devices or Guards', 'Procedures',
        'Personal Protective Equipment', 'Tools Equipment', 'Lifting',
        'Health, Hygiene, Food or Water', 'Safety Signage and Demarcation',
        'Emergency Response', 'Housekeeping', 'Driving/Vehicles',
        'Excavation', 'Line of Fire or Pinch Points', 'Working at Height',
        'Workplace Environment', 'Work Planning & Authorisation', 'Supervision',
        'Situational Awareness', 'Toxic/Flammable Gas', 'Confined Space',
        'Isolation/Lockout', 'Improvement Opportunity', 'Manual/Mechanical Handling',
        'Security', 'Other',
    ]
    severity_unsafe = ['Low', 'Medium', 'High', 'Critical']
    severity_w = [40, 35, 20, 5]
    risk_unsafe = ['Low', 'Medium', 'High']
    risk_w = [40, 40, 20]
    emp_types = ['AON Direct Hire', 'PMC', 'Contractor']

    outcomes = ['Corrected on site', 'Reported to supervisor', 'Work stopped',
                'Immediate rectification', 'Follow-up required', 'Training recommended',
                'Good practice acknowledged', 'Toolbox talk conducted']
    interventions = ['Verbal warning', 'Counseling', 'Re-training', 'Stand-down',
                     'Recognition', 'Positive feedback', 'Safety brief', 'None required']

    today = _date.today()
    count = 0

    for day_offset in range(60):
        d = today - _td(days=day_offset)
        if d.weekday() >= 5:
            continue
        ds = d.isoformat()
        n_obs = _rnd.randint(3, 15)

        for _ in range(n_obs):
            proj = _rnd.choice(projects)
            pid = proj['id'] if isinstance(proj, dict) else proj[0]
            aid = proj['area_id'] if isinstance(proj, dict) else proj[1]
            did = proj['division_id'] if isinstance(proj, dict) else proj[2]

            grp = _rnd.choices(groups, weights=group_weights, k=1)[0]
            is_safe = grp.startswith('Safe')

            sev = 'N/A' if is_safe else _rnd.choices(severity_unsafe, weights=severity_w, k=1)[0]
            risk = 'N/A' if is_safe else _rnd.choices(risk_unsafe, weights=risk_w, k=1)[0]

            observer = _rnd.choice(employees) if employees else None
            obs_name = (observer['name'] if isinstance(observer, dict) else observer[0]) if observer else 'Site Inspector'
            obs_desg = (observer['designation'] if isinstance(observer, dict) else observer[1]) if observer else 'HSE Officer'
            obs_disc = (observer['discipline'] if isinstance(observer, dict) else observer[2]) if observer else 'Safety'
            obs_comp = (observer['contractor_company'] if isinstance(observer, dict) else observer[3]) if observer else 'Main Contractor'

            try:
                db_execute(conn, """INSERT INTO observations
                    (division_id, area_id, project_id, observation_date,
                     observer_name, observer_designation, observer_discipline, observer_company,
                     employee_type, observation_group, observation_type, potential_severity,
                     risk_rating, observation_text, corrective_action, intervention, outcome, remarks)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (did, aid, pid, ds,
                     obs_name, obs_desg, obs_disc, obs_comp,
                     _rnd.choice(emp_types), grp, _rnd.choice(types_list), sev,
                     risk,
                     f"{grp} observation during routine inspection",
                     f"{'No action needed' if is_safe else 'Corrective action taken'}" if _rnd.random() > 0.3 else '',
                     _rnd.choice(interventions),
                     _rnd.choice(outcomes),
                     '' if _rnd.random() > 0.2 else 'Noted during walkthrough'))
                count += 1
            except Exception:
                pass

    try:
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    print(f'  Seeded {count} O&I observations over 60 days')


def _seed_demo_twl(conn):
    """Generate 60 days of TWL readings across areas."""
    import random as _rnd
    from datetime import date as _date, timedelta as _td
    _rnd.seed(55)

    existing = db_fetchone(conn, "SELECT COUNT(*) as c FROM twl_readings")
    cnt = existing['c'] if isinstance(existing, dict) else existing[0]
    if cnt > 20:
        print(f'  TWL already seeded ({cnt} rows), skipping.')
        return

    areas = db_fetchall(conn, "SELECT id FROM areas WHERE active = 1")
    if not areas:
        print('  No areas for TWL seed.')
        return

    sites_by_area = {}
    for area in areas:
        aid = area['id'] if isinstance(area, dict) else area[0]
        s = db_fetchall(conn, """
            SELECT s.id FROM sites s
            JOIN projects p ON p.id = s.project_id
            WHERE p.area_id = ?
        """, (aid,))
        if s:
            sites_by_area[aid] = [x['id'] if isinstance(x, dict) else x[0] for x in s]

    work_types = ['light', 'moderate', 'heavy']
    today = _date.today()
    count = 0

    def _zone(v):
        if v >= 32: return 'low'
        if v >= 28: return 'medium'
        if v >= 25: return 'high'
        return 'extreme'

    for day_offset in range(60):
        d = today - _td(days=day_offset)
        if d.weekday() >= 5:
            continue
        ds = d.isoformat()

        base_temp = _rnd.uniform(32, 48)
        base_humidity = _rnd.uniform(30, 80)
        base_wind = _rnd.uniform(0.5, 6)

        for area in areas:
            aid = area['id'] if isinstance(area, dict) else area[0]
            sid_list = sites_by_area.get(aid)
            sid = _rnd.choice(sid_list) if sid_list else None

            for time_slot in ['06:00', '10:00', '14:00']:
                temp = base_temp + _rnd.uniform(-3, 5) + (4 if '14:' in time_slot else 0)
                hum = base_humidity + _rnd.uniform(-10, 10)
                wind = base_wind + _rnd.uniform(-1, 2)
                twl = _rnd.uniform(22, 36) - (0.15 * (temp - 35)) - (0.05 * (hum - 50))
                twl = max(18, min(38, round(twl, 1)))
                zone = _zone(twl)

                try:
                    db_execute(conn, """INSERT INTO twl_readings
                        (area_id, site_id, reading_date, reading_time, twl_value, risk_zone,
                         temperature, humidity, wind_speed, work_type, notes)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                        (aid, sid, ds, time_slot, twl,
                         zone, round(temp, 1), round(hum, 1), round(wind, 1),
                         _rnd.choice(work_types),
                         f"Routine {time_slot} reading"))
                    count += 1
                except Exception:
                    pass

    try:
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    print(f'  Seeded {count} TWL readings over 60 days')


# Run init_db at import time so gunicorn/production picks it up
# If it fails (e.g. Postgres transaction state), app still starts so / and /reset-admin work
print("Initializing database...")
try:
    init_db()
except Exception as e:
    print(f"  init_db warning: {e}")
    import traceback
    traceback.print_exc()


# ============================================================
# AUTH
# ============================================================

def now_utc():
    return datetime.now(timezone.utc)


def cleanup_expired(conn):
    """Purge expired tokens and 2FA sessions."""
    try:
        db_execute(conn, "DELETE FROM auth_tokens WHERE expires_at < datetime('now')")
        db_execute(conn, "DELETE FROM pending_2fa WHERE expires_at < datetime('now')")
        db_execute(conn, "DELETE FROM download_tokens WHERE expires_at < datetime('now')")
    except Exception:
        pass


def get_current_user():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return None

    conn = get_db()
    try:
        if USE_POSTGRES:
            row = db_fetchone(conn, '''
                SELECT u.* FROM users u
                JOIN auth_tokens t ON t.user_id = u.id
                WHERE t.token = ? AND t.expires_at > NOW()
            ''', (token,))
        else:
            row = db_fetchone(conn, '''
                SELECT u.* FROM users u
                JOIN auth_tokens t ON t.user_id = u.id
                WHERE t.token = ? AND t.expires_at > datetime('now')
            ''', (token,))
        return row
    finally:
        conn.close()


def get_user_projects(conn, user_id, role):
    """Site isolation: each area’s data is separate. Executive and Manager see all; Scanner/Focal Point see only their assigned areas."""
    if role in ('executive', 'admin', 'manager'):
        return None  # see all
    if role == 'focal_point':
        rows = db_fetchall(conn, 'SELECT division_id FROM user_division_access WHERE user_id = ?', (user_id,))
        if not rows:
            return []
        div_ids = [r['division_id'] for r in rows]
        placeholders = ','.join(['?' for _ in div_ids])
        proj_rows = db_fetchall(conn, f'''SELECT p.id FROM projects p
            JOIN areas a ON a.id = p.area_id
            WHERE a.division_id IN ({placeholders}) AND (p.active = 1 OR p.active IS NULL)''', div_ids)
        return [r['id'] for r in proj_rows]
    rows = db_fetchall(conn, 'SELECT project_id FROM user_project_access WHERE user_id = ?', (user_id,))
    return [r['project_id'] for r in rows]


def require_auth(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Authentication required', 'auth_required': True}), 401
        request.user = user
        return f(*args, **kwargs)
    return decorated


def require_role(*roles):
    def decorator(f):
        @functools.wraps(f)
        def decorated(*args, **kwargs):
            user = get_current_user()
            if not user:
                return jsonify({'error': 'Authentication required', 'auth_required': True}), 401
            if user['role'] not in roles:
                return jsonify({'error': f'Requires role: {", ".join(roles)}'}), 403
            request.user = user
            return f(*args, **kwargs)
        return decorated
    return decorator


def audit(user_id, action, detail=''):
    try:
        conn = get_db()
        try:
            db_execute(conn, 'INSERT INTO audit_log (user_id, action, detail, ip_address) VALUES (?, ?, ?, ?)',
                       (user_id, action, detail, request.remote_addr))
            conn.commit()
        finally:
            conn.close()
    except Exception:
        pass


def check_ip_rate_limit(ip):
    now = datetime.now().timestamp()
    if ip not in LOGIN_ATTEMPTS:
        LOGIN_ATTEMPTS[ip] = []
    LOGIN_ATTEMPTS[ip] = [t for t in LOGIN_ATTEMPTS[ip] if now - t < LOGIN_WINDOW]
    if len(LOGIN_ATTEMPTS[ip]) >= LOGIN_MAX_PER_IP:
        return False
    LOGIN_ATTEMPTS[ip].append(now)
    return True


def apply_project_filter(conn, query, params, user, table_alias='a'):
    allowed = get_user_projects(conn, user['id'], user['role'])
    if allowed is not None:
        if not allowed:
            query += f' AND {table_alias}.project_id = -1'
        elif len(allowed) == 1:
            query += f' AND {table_alias}.project_id = ?'
            params.append(allowed[0])
        else:
            placeholders = ','.join(['?' for _ in allowed])
            query += f' AND {table_alias}.project_id IN ({placeholders})'
            params.extend(allowed)
    return query, params


def check_project_access(conn, user, project_id):
    allowed = get_user_projects(conn, user['id'], user['role'])
    if allowed is None:
        return True
    return int(project_id) in allowed


def get_user_divisions(conn, user_id, role):
    """Division IDs the user can access. Focal point: from user_division_access; others: all (None)."""
    if role != 'focal_point':
        return None
    rows = db_fetchall(conn, 'SELECT division_id FROM user_division_access WHERE user_id = ?', (user_id,))
    return [r['division_id'] for r in rows]


# ============================================================
# AUTH ROUTES
# ============================================================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Two login paths: scanner (username+PIN) and admin (username+2FA)."""
    try:
        return _do_login()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500

def _do_login():
    if not check_ip_rate_limit(request.remote_addr):
        return jsonify({'success': False, 'message': 'Too many attempts. Wait 1 minute.'}), 429

    data = request.json or {}
    login_mode = (data.get('login_mode') or 'admin').strip()
    username = (data.get('username') or '').strip().lower()

    if not username:
        return jsonify({'success': False, 'message': 'Username required'}), 400

    conn = get_db()
    try:
        cleanup_expired(conn)
        user = db_fetchone(conn, 'SELECT * FROM users WHERE username = ?', (username,))
        if not user:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

        if login_mode == 'scanner':
            pin = (data.get('pin') or '').strip()
            if not pin:
                return jsonify({'success': False, 'message': 'PIN required'}), 400
            if user['role'] not in ('scanner',):
                return jsonify({'success': False, 'message': 'Use Admin sign-in for this account'}), 403
            if not verify_pin(pin, user['pin_hash']):
                audit(user['id'], 'login_failed', 'bad PIN')
                return jsonify({'success': False, 'message': 'Invalid PIN'}), 401
        else:
            totp_code = (data.get('totp_code') or '').strip()
            if not totp_code:
                return jsonify({'success': False, 'message': '2FA code required'}), 400
            if len(totp_code) != 6 or not totp_code.isdigit():
                return jsonify({'success': False, 'message': 'Enter the 6-digit code from your authenticator app'}), 400
            if user['role'] == 'scanner':
                return jsonify({'success': False, 'message': 'Use Scanner sign-in for this account'}), 403
            secret = (user.get('totp_secret') or '').strip()
            enabled = user.get('totp_enabled')
            if not secret or enabled in (None, 0, False, '0'):
                return jsonify({'success': False, 'message': '2FA not set up. Contact your administrator.'}), 403
            try:
                totp = pyotp.TOTP(secret)
                if not totp.verify(totp_code, valid_window=2):
                    audit(user['id'], '2fa_failed', request.remote_addr)
                    return jsonify({'success': False, 'message': 'Invalid 2FA code. Try the newest code from your app.'}), 401
            except Exception:
                return jsonify({'success': False, 'message': '2FA verification error. Contact support.'}), 500

        token = secrets.token_urlsafe(48)
        device = request.headers.get('User-Agent', '')[:200]
        db_execute(conn, "INSERT INTO auth_tokens (user_id, token, device_info, expires_at) VALUES (?, ?, ?, datetime('now', '+24 hours'))",
                   (user['id'], token, device))
        db_execute(conn, "UPDATE users SET last_login = datetime('now') WHERE id = ?", (user['id'],))
        conn.commit()

        allowed = get_user_projects(conn, user['id'], user['role'])
        allowed_divs = get_user_divisions(conn, user['id'], user['role'])
        audit(user['id'], 'login', f'{login_mode} | {device[:60]}')

        return jsonify({
            'success': True,
            'token': token,
            'user': {
                'id': user['id'], 'username': user['username'],
                'display_name': user['display_name'], 'role': user['role'],
                'email': user.get('email') or '', 'designation': user.get('designation') or '',
                'must_change_pin': False,
                'totp_enabled': bool(user.get('totp_enabled')),
                'allowed_projects': allowed,
                'allowed_divisions': allowed_divs
            }
        })
    finally:
        conn.close()


@app.route('/api/auth/verify-2fa', methods=['POST'])
def verify_2fa():
    data = request.json
    pending_token = data.get('pending_token', '').strip()
    totp_code = data.get('totp_code', '').strip()

    if not pending_token or not totp_code:
        return jsonify({'success': False, 'message': '2FA code required'}), 400

    conn = get_db()
    try:
        if USE_POSTGRES:
            pending = db_fetchone(conn, '''
                SELECT p.user_id, p.pending_token, u.* FROM pending_2fa p
                JOIN users u ON u.id = p.user_id WHERE p.pending_token = ? AND p.expires_at > NOW()
            ''', (pending_token,))
        else:
            pending = db_fetchone(conn, '''
                SELECT p.user_id, p.pending_token, u.* FROM pending_2fa p
                JOIN users u ON u.id = p.user_id WHERE p.pending_token = ? AND p.expires_at > datetime('now')
            ''', (pending_token,))

        if not pending:
            return jsonify({'success': False, 'message': 'Session expired, login again'}), 401

        totp = pyotp.TOTP(pending['totp_secret'])
        if not totp.verify(totp_code, valid_window=1):
            db_execute(conn, 'DELETE FROM pending_2fa WHERE pending_token = ?', (pending_token,))
            conn.commit()
            audit(pending['user_id'], '2fa_failed')
            return jsonify({'success': False, 'message': 'Invalid code. Login again.'}), 401

        db_execute(conn, 'DELETE FROM pending_2fa WHERE user_id = ?', (pending['user_id'],))
        token = secrets.token_urlsafe(48)
        device = request.headers.get('User-Agent', '')[:200]
        db_execute(conn, "INSERT INTO auth_tokens (user_id, token, device_info, expires_at) VALUES (?, ?, ?, datetime('now', '+24 hours'))",
                   (pending['user_id'], token, device))
        db_execute(conn, "UPDATE users SET last_login = datetime('now') WHERE id = ?", (pending['user_id'],))
        conn.commit()

        allowed = get_user_projects(conn, pending['user_id'], pending['role'])
        allowed_divs = get_user_divisions(conn, pending['user_id'], pending['role'])
        audit(pending['user_id'], 'login_2fa')

        return jsonify({
            'success': True, 'token': token,
            'user': {
                'id': pending['user_id'], 'username': pending['username'],
                'display_name': pending['display_name'], 'role': pending['role'],
                'email': pending.get('email') or '', 'designation': pending.get('designation') or '',
                'must_change_pin': bool(pending.get('must_change_pin')),
                'totp_enabled': True,
                'allowed_projects': allowed,
                'allowed_divisions': allowed_divs
            }
        })
    finally:
        conn.close()


@app.route('/api/auth/logout', methods=['POST'])
@require_auth
def logout():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    conn = get_db()
    try:
        db_execute(conn, 'DELETE FROM auth_tokens WHERE token = ?', (token,))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'logout')
    return jsonify({'success': True})


@app.route('/api/auth/change-pin', methods=['POST'])
@require_auth
def change_pin():
    data = request.json
    new_pin = data.get('new_pin', '').strip()
    current_pin = data.get('current_pin', '').strip()

    if not new_pin or len(new_pin) < 4 or len(new_pin) > 8 or not new_pin.isdigit():
        return jsonify({'success': False, 'message': 'PIN must be 4-8 digits'}), 400
    if is_weak_pin(new_pin):
        return jsonify({'success': False, 'message': 'PIN too simple. Avoid sequences and repeated digits.'}), 400

    conn = get_db()
    try:
        user = db_fetchone(conn, 'SELECT * FROM users WHERE id = ?', (request.user['id'],))
        if not user.get('must_change_pin') and not verify_pin(current_pin, user['pin_hash']):
            return jsonify({'success': False, 'message': 'Current PIN incorrect'}), 401

        db_execute(conn, 'UPDATE users SET pin_hash = ?, must_change_pin = 0 WHERE id = ?',
                   (hash_pin(new_pin), request.user['id']))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'pin_changed')
    return jsonify({'success': True, 'message': 'PIN updated'})


@app.route('/api/auth/me')
@require_auth
def auth_me():
    conn = get_db()
    try:
        allowed = get_user_projects(conn, request.user['id'], request.user['role'])
        allowed_divs = get_user_divisions(conn, request.user['id'], request.user['role'])
    finally:
        conn.close()
    return jsonify({
        'id': request.user['id'], 'username': request.user['username'],
        'display_name': request.user['display_name'], 'role': request.user['role'],
        'email': request.user.get('email') or '', 'designation': request.user.get('designation') or '',
        'must_change_pin': bool(request.user.get('must_change_pin')),
        'totp_enabled': bool(request.user.get('totp_enabled')),
        'allowed_projects': allowed,
        'allowed_divisions': allowed_divs
    })


# --- 2FA Setup ---

@app.route('/api/auth/2fa/setup', methods=['POST'])
@require_auth
def setup_2fa():
    conn = get_db()
    try:
        user = db_fetchone(conn, 'SELECT * FROM users WHERE id = ?', (request.user['id'],))
        if user.get('totp_enabled'):
            return jsonify({'success': False, 'message': '2FA already enabled'}), 400

        secret = pyotp.random_base32()
        db_execute(conn, 'UPDATE users SET totp_secret = ? WHERE id = ?', (secret, request.user['id']))
        conn.commit()
    finally:
        conn.close()

    uri = pyotp.TOTP(secret).provisioning_uri(name=user['username'], issuer_name='PTC POB Tracker')
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=6, border=3)
    qr.add_data(uri)
    qr.make(fit=True)
    buf = BytesIO()
    qr.make_image(fill_color="black", back_color="white").save(buf, format='PNG')
    return jsonify({'success': True, 'secret': secret, 'qr_code': base64.b64encode(buf.getvalue()).decode(), 'uri': uri})


@app.route('/api/auth/2fa/confirm', methods=['POST'])
@require_auth
def confirm_2fa():
    totp_code = request.json.get('totp_code', '').strip()
    conn = get_db()
    try:
        user = db_fetchone(conn, 'SELECT * FROM users WHERE id = ?', (request.user['id'],))
        if not user.get('totp_secret'):
            return jsonify({'success': False, 'message': 'Run setup first'}), 400
        if not pyotp.TOTP(user['totp_secret']).verify(totp_code, valid_window=1):
            return jsonify({'success': False, 'message': 'Invalid code'}), 400
        db_execute(conn, 'UPDATE users SET totp_enabled = 1 WHERE id = ?', (request.user['id'],))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], '2fa_enabled')
    return jsonify({'success': True, 'message': '2FA enabled'})


@app.route('/api/auth/2fa/disable', methods=['POST'])
@require_auth
def disable_2fa():
    """Disable 2FA; requires current 2FA code (sign-in is 2FA only)."""
    data = request.json or {}
    totp_code = (data.get('totp_code') or data.get('pin') or '').strip()
    conn = get_db()
    try:
        user = db_fetchone(conn, 'SELECT * FROM users WHERE id = ?', (request.user['id'],))
        if not user.get('totp_secret'):
            return jsonify({'success': False, 'message': '2FA is not enabled'}), 400
        if not totp_code or len(totp_code) != 6:
            return jsonify({'success': False, 'message': 'Enter your current 6-digit 2FA code to disable'}), 400
        if not pyotp.TOTP(user['totp_secret']).verify(totp_code, valid_window=1):
            return jsonify({'success': False, 'message': 'Incorrect 2FA code'}), 401
        db_execute(conn, 'UPDATE users SET totp_enabled = 0, totp_secret = NULL WHERE id = ?', (request.user['id'],))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], '2fa_disabled')
    return jsonify({'success': True, 'message': '2FA disabled'})


# ============================================================
# EMAIL NOTIFICATIONS
# ============================================================

def _generate_qr_image_bytes(uri):
    """Generate a QR code PNG as bytes for embedding in emails."""
    img = qrcode.make(uri, box_size=6, border=2)
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.read()


def send_welcome_email(recipient_email, display_name, username, role, pin=None,
                       totp_secret=None, totp_uri=None):
    """Send account-creation notification email in a background thread."""
    if not SMTP_PASS or not recipient_email:
        return

    def _send():
        try:
            msg = MIMEMultipart('related')
            msg['From'] = f'PTC POB Tracker <{SMTP_USER}>'
            msg['To'] = recipient_email
            msg['Subject'] = 'Your PTC POB Tracker Account Has Been Created'

            role_labels = {
                'scanner': 'Contractor Scanner',
                'focal_point': 'Divisional Focal Point',
                'manager': 'HSE Manager',
                'admin': 'Administrator',
                'executive': 'Executive',
            }
            role_label = role_labels.get(role, role.title())
            app_url = 'https://ptc-pob-tracker.onrender.com'

            if role == 'scanner':
                credentials_html = f'''
                <tr><td style="padding:8px 12px;font-weight:600;color:#555;width:160px">Username</td>
                    <td style="padding:8px 12px;font-family:monospace;font-size:15px">{username}</td></tr>
                <tr><td style="padding:8px 12px;font-weight:600;color:#555">PIN</td>
                    <td style="padding:8px 12px;font-family:monospace;font-size:15px">{pin or '(provided separately)'}</td></tr>
                '''
                auth_instructions = '''
                <p style="color:#555;line-height:1.6">
                    Use the <strong>Contractor Scanner</strong> login on the app. Enter your username and PIN to sign in.
                </p>'''
            else:
                credentials_html = f'''
                <tr><td style="padding:8px 12px;font-weight:600;color:#555;width:160px">Username</td>
                    <td style="padding:8px 12px;font-family:monospace;font-size:15px">{username}</td></tr>
                <tr><td style="padding:8px 12px;font-weight:600;color:#555">2FA Setup Key</td>
                    <td style="padding:8px 12px;font-family:monospace;font-size:13px;word-break:break-all">{totp_secret}</td></tr>
                '''
                auth_instructions = '''
                <p style="color:#555;line-height:1.6">
                    Use the <strong>Admin / Manager</strong> login on the app. You will need a 2FA code from your authenticator app to sign in.
                </p>
                <p style="color:#555;line-height:1.6">
                    <strong>Setup your authenticator:</strong> Open Google Authenticator, Microsoft Authenticator, or any TOTP app and scan the QR code below:
                </p>
                <div style="text-align:center;margin:16px 0">
                    <img src="cid:qrcode" alt="2FA QR Code" style="border:1px solid #ddd;border-radius:8px;padding:8px;background:#fff" />
                </div>
                <p style="color:#888;font-size:12px;text-align:center">
                    Can&rsquo;t scan? Enter the setup key manually in your authenticator app.
                </p>'''

            html = f'''
            <div style="max-width:560px;margin:0 auto;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif">
                <div style="background:#0d47a1;color:#fff;padding:20px 24px;border-radius:8px 8px 0 0">
                    <h2 style="margin:0;font-size:18px">PTC POB Tracker</h2>
                </div>
                <div style="background:#ffffff;padding:24px;border:1px solid #e0e0e0;border-top:none;border-radius:0 0 8px 8px">
                    <p style="color:#333;font-size:15px;line-height:1.6">Dear <strong>{display_name}</strong>,</p>
                    <p style="color:#555;line-height:1.6">
                        Your account on the <strong>PTC POB Tracker</strong> has been created. Below are your login details:
                    </p>
                    <table style="width:100%;border-collapse:collapse;margin:16px 0;background:#f8f9fa;border-radius:6px;overflow:hidden">
                        <tr><td style="padding:8px 12px;font-weight:600;color:#555;width:160px">Role</td>
                            <td style="padding:8px 12px">{role_label}</td></tr>
                        {credentials_html}
                        <tr><td style="padding:8px 12px;font-weight:600;color:#555">App URL</td>
                            <td style="padding:8px 12px"><a href="{app_url}" style="color:#0d47a1">{app_url}</a></td></tr>
                    </table>
                    {auth_instructions}
                    <hr style="border:none;border-top:1px solid #eee;margin:24px 0" />
                    <p style="color:#999;font-size:12px;line-height:1.5">
                        This is an automated message. If you did not expect this, please contact your divisional focal point or HSE manager.
                    </p>
                </div>
            </div>'''

            msg.attach(MIMEText(html, 'html'))

            if totp_uri and role != 'scanner':
                qr_bytes = _generate_qr_image_bytes(totp_uri)
                qr_img = MIMEImage(qr_bytes, _subtype='png')
                qr_img.add_header('Content-ID', '<qrcode>')
                qr_img.add_header('Content-Disposition', 'inline', filename='2fa-qrcode.png')
                msg.attach(qr_img)

            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.send_message(msg)
            print(f"[EMAIL] Welcome email sent to {recipient_email}")
        except Exception as e:
            print(f"[EMAIL ERROR] Failed to send to {recipient_email}: {e}")

    threading.Thread(target=_send, daemon=True).start()


# ============================================================
# USER MANAGEMENT
# ============================================================

@app.route('/api/users')
@require_role('executive', 'admin')
def api_users():
    conn = get_db()
    try:
        rows = db_fetchall(conn, '''
            SELECT id, username, display_name, email, designation, role, active, must_change_pin,
                   failed_attempts, locked_until, totp_enabled, created_at, last_login
            FROM users ORDER BY role, display_name
        ''')
        for row in rows:
            access = db_fetchall(conn, '''
                SELECT p.id, p.name FROM user_project_access upa
                JOIN projects p ON p.id = upa.project_id WHERE upa.user_id = ?
            ''', (row['id'],))
            row['projects'] = access
            divs = db_fetchall(conn, '''
                SELECT d.id, d.name FROM user_division_access uda
                JOIN divisions d ON d.id = uda.division_id WHERE uda.user_id = ?
            ''', (row['id'],))
            row['divisions'] = divs
    finally:
        conn.close()
    return jsonify(rows)


@app.route('/api/users', methods=['POST'])
@require_role('executive', 'admin')
def api_add_user():
    data = request.json
    username = data.get('username', '').strip().lower()
    display_name = data.get('display_name', '').strip()
    role = data.get('role', 'supervisor')
    pin = data.get('pin', '1234').strip() or '1234'
    project_ids = data.get('project_ids', [])
    division_ids = data.get('division_ids', [])

    if not username or not display_name:
        return jsonify({'success': False, 'message': 'Username and name required'}), 400
    if role not in ROLE_HIERARCHY:
        return jsonify({'success': False, 'message': f'Invalid role. Use: {", ".join(ROLE_HIERARCHY.keys())}'}), 400
    if request.user['role'] != 'executive' and role in ('executive', 'admin'):
        return jsonify({'success': False, 'message': 'Only executives can create admin/executive users'}), 403

    email = data.get('email', '').strip()
    designation = data.get('designation', '').strip()
    if role in ('scanner', 'focal_point') and not designation:
        return jsonify({'success': False, 'message': 'Designation required for scanner/focal point'}), 400
    if role == 'scanner' and not project_ids:
        return jsonify({'success': False, 'message': 'Assign at least one project for scanner'}), 400
    if role == 'focal_point' and not division_ids:
        return jsonify({'success': False, 'message': 'Assign at least one division for divisional focal point'}), 400

    needs_2fa = role != 'scanner'
    totp_secret = pyotp.random_base32() if needs_2fa else None
    conn = get_db()
    try:
        db_execute(conn, '''
            INSERT INTO users (username, display_name, email, designation, pin_hash, role, must_change_pin, totp_secret, totp_enabled)
            VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?)
        ''', (username, display_name, email or None, designation or None, hash_pin(pin), role,
              totp_secret, 1 if needs_2fa else 0))
        conn.commit()

        user = db_fetchone(conn, 'SELECT id FROM users WHERE username = ?', (username,))
        if project_ids and role == 'scanner':
            for pid in project_ids:
                try:
                    db_execute(conn, 'INSERT INTO user_project_access (user_id, project_id) VALUES (?, ?)',
                               (user['id'], pid))
                except Exception:
                    pass
            conn.commit()
        if division_ids and role == 'focal_point':
            for did in division_ids:
                try:
                    db_execute(conn, 'INSERT INTO user_division_access (user_id, division_id) VALUES (?, ?)',
                               (user['id'], did))
                except Exception:
                    pass
            conn.commit()

        audit(request.user['id'], 'user_created', f'{username} ({role})')
        result = {'success': True}
        totp_uri = None
        if needs_2fa:
            totp_uri = pyotp.TOTP(totp_secret).provisioning_uri(name=username, issuer_name='PTC POB Tracker')
            result['message'] = f'User {username} created. Give them the 2FA setup below to sign in.'
            result['totp_setup'] = {'secret': totp_secret, 'uri': totp_uri}
        else:
            result['message'] = f'Scanner {username} created with PIN {pin}. Share the username and PIN with them.'

        if email:
            send_welcome_email(
                recipient_email=email,
                display_name=display_name,
                username=username,
                role=role,
                pin=pin if role == 'scanner' else None,
                totp_secret=totp_secret,
                totp_uri=totp_uri,
            )
            result['email_sent'] = True
            result['message'] += f' A welcome email has been sent to {email}.'

        return jsonify(result)
    except Exception as e:
        if 'UNIQUE' in str(e).upper() or 'unique' in str(e).lower():
            return jsonify({'success': False, 'message': 'Username already exists'}), 400
        return jsonify({'success': False, 'message': str(e)}), 400
    finally:
        conn.close()


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@require_role('executive', 'admin')
def api_update_user(user_id):
    data = request.json
    conn = get_db()
    try:
        if 'role' in data and data['role'] in ROLE_HIERARCHY:
            db_execute(conn, 'UPDATE users SET role = ? WHERE id = ?', (data['role'], user_id))
        if 'display_name' in data:
            db_execute(conn, 'UPDATE users SET display_name = ? WHERE id = ?', (data['display_name'], user_id))
        if 'email' in data:
            db_execute(conn, 'UPDATE users SET email = ? WHERE id = ?', (data['email'], user_id))
        if 'designation' in data:
            db_execute(conn, 'UPDATE users SET designation = ? WHERE id = ?', (data['designation'], user_id))
        if 'project_ids' in data:
            db_execute(conn, 'DELETE FROM user_project_access WHERE user_id = ?', (user_id,))
            for pid in data['project_ids']:
                try:
                    db_execute(conn, 'INSERT INTO user_project_access (user_id, project_id) VALUES (?, ?)', (user_id, pid))
                except Exception:
                    pass
        if 'division_ids' in data:
            db_execute(conn, 'DELETE FROM user_division_access WHERE user_id = ?', (user_id,))
            for did in data['division_ids']:
                try:
                    db_execute(conn, 'INSERT INTO user_division_access (user_id, division_id) VALUES (?, ?)', (user_id, did))
                except Exception:
                    pass
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'user_updated', f'#{user_id}')
    return jsonify({'success': True})


@app.route('/api/users/<int:user_id>/reset-pin', methods=['POST'])
@require_role('executive', 'admin')
def api_reset_pin(user_id):
    new_pin = request.json.get('pin', '1234')
    conn = get_db()
    try:
        db_execute(conn, 'UPDATE users SET pin_hash = ?, must_change_pin = 1, failed_attempts = 0, locked_until = NULL WHERE id = ?',
                   (hash_pin(new_pin), user_id))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'pin_reset', f'#{user_id}')
    return jsonify({'success': True, 'message': f'PIN reset to {new_pin}'})


@app.route('/api/users/<int:user_id>/reset-2fa', methods=['POST'])
@require_role('executive', 'admin')
def admin_reset_2fa(user_id):
    """Generate new 2FA secret for user; return secret/uri so admin can give it to the user (sign-in is 2FA only)."""
    conn = get_db()
    try:
        row = db_fetchone(conn, 'SELECT username FROM users WHERE id = ?', (user_id,))
        if not row:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        new_secret = pyotp.random_base32()
        db_execute(conn, 'UPDATE users SET totp_secret = ?, totp_enabled = 1 WHERE id = ?', (new_secret, user_id))
        conn.commit()
        uri = pyotp.TOTP(new_secret).provisioning_uri(name=row['username'], issuer_name='PTC POB Tracker')
        audit(request.user['id'], '2fa_reset', f'#{user_id}')
        return jsonify({
            'success': True,
            'message': 'New 2FA secret generated. Give the user this setup to sign in.',
            'totp_setup': {'secret': new_secret, 'uri': uri}
        })
    finally:
        conn.close()


# ============================================================
# DIVISIONS / AREAS / PROJECTS (filtered by user access)
# ============================================================

@app.route('/api/ping')
def ping():
    return jsonify({'pong': True, 'db_type': 'PostgreSQL' if DATABASE_URL else 'SQLite'})

@app.route('/api/health')
def health():
    """Diagnostic endpoint."""
    try:
        conn = get_db()
        try:
            row = db_fetchone(conn, 'SELECT COUNT(*) as c FROM users')
            users = row['c'] if isinstance(row, dict) else row[0]
            prow = db_fetchone(conn, 'SELECT COUNT(*) as c FROM projects')
            projects = prow['c'] if isinstance(prow, dict) else prow[0]
            dem_row = db_fetchone(conn, "SELECT COUNT(*) as c FROM employees WHERE employee_no LIKE ?", ('DEM-%',))
            dem = dem_row['c'] if isinstance(dem_row, dict) else dem_row[0]
            total_row = db_fetchone(conn, "SELECT COUNT(*) as c FROM employees")
            total_emp = total_row['c'] if isinstance(total_row, dict) else total_row[0]
            att_row = db_fetchone(conn, "SELECT COUNT(*) as c FROM attendance")
            att = att_row['c'] if isinstance(att_row, dict) else att_row[0]
            obs_row = db_fetchone(conn, "SELECT COUNT(*) as c FROM observations")
            obs = obs_row['c'] if isinstance(obs_row, dict) else obs_row[0]
            twl_row = db_fetchone(conn, "SELECT COUNT(*) as c FROM twl_readings")
            twl = twl_row['c'] if isinstance(twl_row, dict) else twl_row[0]
            sample = db_fetchone(conn, "SELECT employee_no, name FROM employees LIMIT 1")
            sample_data = dict(sample) if sample else None
            return jsonify({
                'status': 'ok', 'users': users, 'projects': projects,
                'db': 'PostgreSQL' if DATABASE_URL else 'SQLite',
                'total_employees': total_emp, 'dem_employees': dem,
                'attendance': att, 'observations': obs, 'twl_readings': twl,
                'sample': sample_data
            })
        finally:
            conn.close()
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/')
def index():
    return send_from_directory('public', 'index.html')


@app.route('/api/divisions')
@require_auth
def api_divisions():
    """List divisions. Focal point sees only their assigned divisions; admin/manager sees all."""
    conn = get_db()
    try:
        if request.user['role'] == 'focal_point':
            rows = db_fetchall(conn, '''SELECT d.id, d.name, d.active
                FROM divisions d
                JOIN user_division_access uda ON uda.division_id = d.id
                WHERE uda.user_id = ? AND (d.active = 1 OR d.active IS NULL)
                ORDER BY d.name''', (request.user['id'],))
        else:
            rows = db_fetchall(conn, 'SELECT id, name, active FROM divisions WHERE active = 1 OR active IS NULL ORDER BY name')
    finally:
        conn.close()
    return jsonify(rows)


@app.route('/api/areas')
@require_auth
def api_areas():
    """List areas. Optional division_id= to filter. Respects user access."""
    division_id = request.args.get('division_id', type=int)
    conn = get_db()
    try:
        query = '''SELECT a.id, a.division_id, a.name, a.active, d.name as division_name
            FROM areas a
            JOIN divisions d ON d.id = a.division_id
            WHERE (a.active = 1 OR a.active IS NULL)'''
        params = []
        if division_id:
            query += ' AND a.division_id = ?'
            params.append(division_id)
        if request.user['role'] == 'focal_point':
            query += ' AND EXISTS (SELECT 1 FROM user_division_access uda WHERE uda.user_id = ? AND uda.division_id = a.division_id)'
            params.append(request.user['id'])
        query += ' ORDER BY a.name'
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    return jsonify(rows)


@app.route('/api/divisions', methods=['POST'])
@require_role('executive', 'admin')
def api_add_division():
    name = (request.json or {}).get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Name required'}), 400
    conn = get_db()
    try:
        existing = db_fetchone(conn, 'SELECT id FROM divisions WHERE name = ?', (name,))
        if existing:
            return jsonify({'success': False, 'message': 'Division already exists'}), 409
        db_execute(conn, 'INSERT INTO divisions (name, active) VALUES (?, 1)', (name,))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'division_added', name)
    return jsonify({'success': True})


@app.route('/api/divisions/<int:div_id>', methods=['PUT'])
@require_role('executive', 'admin')
def api_update_division(div_id):
    name = (request.json or {}).get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Name required'}), 400
    conn = get_db()
    try:
        db_execute(conn, 'UPDATE divisions SET name = ? WHERE id = ?', (name, div_id))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'division_renamed', f'#{div_id} -> {name}')
    return jsonify({'success': True})


@app.route('/api/divisions/<int:div_id>', methods=['DELETE'])
@require_role('executive', 'admin')
def api_delete_division(div_id):
    conn = get_db()
    try:
        db_execute(conn, 'UPDATE divisions SET active = 0 WHERE id = ?', (div_id,))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'division_deleted', f'#{div_id}')
    return jsonify({'success': True})


@app.route('/api/areas', methods=['POST'])
@require_role('executive', 'admin')
def api_add_area():
    data = request.json or {}
    name = data.get('name', '').strip()
    division_id = data.get('division_id')
    if not name or not division_id:
        return jsonify({'success': False, 'message': 'Name and division_id required'}), 400
    conn = get_db()
    try:
        existing = db_fetchone(conn, 'SELECT id FROM areas WHERE division_id = ? AND name = ?', (division_id, name))
        if existing:
            return jsonify({'success': False, 'message': 'Area already exists in this division'}), 409
        db_execute(conn, 'INSERT INTO areas (division_id, name, active) VALUES (?, ?, 1)', (division_id, name))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'area_added', name)
    return jsonify({'success': True})


@app.route('/api/areas/<int:area_id>', methods=['PUT'])
@require_role('executive', 'admin')
def api_update_area(area_id):
    name = (request.json or {}).get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Name required'}), 400
    conn = get_db()
    try:
        db_execute(conn, 'UPDATE areas SET name = ? WHERE id = ?', (name, area_id))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'area_renamed', f'#{area_id} -> {name}')
    return jsonify({'success': True})


@app.route('/api/areas/<int:area_id>', methods=['DELETE'])
@require_role('executive', 'admin')
def api_delete_area(area_id):
    conn = get_db()
    try:
        db_execute(conn, 'UPDATE areas SET active = 0 WHERE id = ?', (area_id,))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'area_deleted', f'#{area_id}')
    return jsonify({'success': True})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@require_role('executive', 'admin')
def api_delete_user(user_id):
    if user_id == request.user['id']:
        return jsonify({'success': False, 'message': 'Cannot delete yourself'}), 400
    conn = get_db()
    try:
        db_execute(conn, 'DELETE FROM user_project_access WHERE user_id = ?', (user_id,))
        db_execute(conn, 'DELETE FROM user_division_access WHERE user_id = ?', (user_id,))
        db_execute(conn, 'DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'user_deleted', f'#{user_id}')
    return jsonify({'success': True})


@app.route('/api/projects')
@require_auth
def api_projects():
    """List projects with division/area info. Optional area_id= or division_id=. Filtered by role."""
    area_id = request.args.get('area_id', type=int)
    division_id = request.args.get('division_id', type=int)
    conn = get_db()
    try:
        query = '''SELECT p.id, p.area_id, p.name, p.description, p.agreement_no,
            p.contract_number, p.contractor_company, p.active,
            a.name as area_name, a.division_id, d.name as division_name,
            COUNT(DISTINCT e.id) as employee_count, COUNT(DISTINCT s.id) as site_count
            FROM projects p
            LEFT JOIN areas a ON a.id = p.area_id
            LEFT JOIN divisions d ON d.id = a.division_id
            LEFT JOIN employees e ON e.project_id = p.id AND e.active = 1
            LEFT JOIN sites s ON s.project_id = p.id AND s.active = 1
            WHERE (p.active = 1 OR p.active IS NULL)'''
        params = []
        if area_id:
            query += ' AND p.area_id = ?'
            params.append(area_id)
        if division_id:
            query += ' AND a.division_id = ?'
            params.append(division_id)
        allowed = get_user_projects(conn, request.user['id'], request.user['role'])
        if allowed is not None:
            if not allowed:
                query += ' AND p.id = -1'
            else:
                placeholders = ','.join(['?' for _ in allowed])
                query += f' AND p.id IN ({placeholders})'
                params.extend(allowed)
        query += " GROUP BY p.id, p.area_id, p.name, p.contract_number, p.contractor_company, a.id, a.name, a.division_id, d.id, d.name ORDER BY COALESCE(d.name, ''), COALESCE(a.name, ''), p.name"
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    return jsonify(rows)


@app.route('/api/projects', methods=['POST'])
@require_role('executive', 'admin', 'focal_point')
def api_add_project():
    data = request.json or {}
    area_id = data.get('area_id')
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Project name required'}), 400
    if not area_id and request.user['role'] != 'executive' and request.user['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Area required'}), 400
    conn = get_db()
    try:
        db_execute(conn, 'INSERT INTO projects (area_id, name, description, agreement_no, active) VALUES (?, ?, ?, ?, 1)',
                   (area_id, name, data.get('description', ''), data.get('agreement_no', '')))
        conn.commit()
        proj = db_fetchone(conn, 'SELECT * FROM projects WHERE area_id = ? AND name = ?', (area_id, name))
        if proj:
            db_execute(conn, 'INSERT INTO sites (name, project_id, active) VALUES (?, ?, 1)', (name, proj['id']))
            conn.commit()
        audit(request.user['id'], 'project_created', name)
        return jsonify({'success': True, 'project': proj})
    except Exception as e:
        if 'UNIQUE' in str(e).upper():
            return jsonify({'success': False, 'message': 'Project already exists in this area'}), 400
        return jsonify({'success': False, 'message': str(e)}), 400
    finally:
        conn.close()


@app.route('/api/sites')
@require_auth
def api_sites():
    project_id = request.args.get('project_id')
    conn = get_db()
    try:
        query = 'SELECT s.* FROM sites s WHERE s.active = 1'
        params = []
        if project_id:
            query += ' AND s.project_id = ?'
            params.append(project_id)
        query, params = apply_project_filter(conn, query, params, request.user, 's')
        query += ' ORDER BY s.name'
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    return jsonify(rows)


@app.route('/api/sites', methods=['POST'])
@require_role('executive', 'admin')
def api_add_site():
    data = request.json
    conn = get_db()
    try:
        db_execute(conn, 'INSERT INTO sites (name, project_id, description) VALUES (?, ?, ?)',
                   (data['name'], data['project_id'], data.get('description', '')))
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        return jsonify({'success': False, 'message': 'Site already exists'}), 400
    finally:
        conn.close()


# ============================================================
# IMPORT
# ============================================================

@app.route('/api/import-roster', methods=['POST'])
@require_role('executive', 'admin')
def api_import_roster():
    project_name = request.json.get('project_name') if request.is_json else None
    count, msg = import_roster(project_name=project_name)
    audit(request.user['id'], 'roster_imported', msg)
    return jsonify({'success': True, 'count': count, 'message': msg})


@app.route('/api/import-csv', methods=['POST'])
@require_role('executive', 'admin')
def api_import_csv():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file'}), 400
    f = request.files['file']
    project_name = request.form.get('project_name', '').strip()
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.csv', delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name
    try:
        count, msg = import_roster(tmp_path, project_name=project_name or None)
    finally:
        os.unlink(tmp_path)
    audit(request.user['id'], 'csv_imported', msg)
    return jsonify({'success': True, 'count': count, 'message': msg})


@app.route('/api/import-excel', methods=['POST'])
@require_role('executive', 'admin')
def api_import_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Upload an .xlsx file'}), 400
    area_id = request.form.get('area_id', type=int)
    if not area_id:
        return jsonify({'success': False, 'message': 'Select an area for the import'}), 400
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name
    try:
        count, msg = import_excel(tmp_path, area_id=area_id)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    audit(request.user['id'], 'excel_imported', msg)
    return jsonify({'success': True, 'count': count, 'message': msg})


# ============================================================
# EMPLOYEES / SCAN / HEADCOUNT (project-scoped)
# ============================================================

@app.route('/api/contractors')
@require_auth
def api_contractors():
    """Return distinct contractor company names (from projects + employee subcontractor)."""
    project_id = request.args.get('project_id', '')
    area_id = request.args.get('area_id', type=int)
    division_id = request.args.get('division_id', type=int)
    conn = get_db()
    try:
        names = set()
        # From projects table
        pq = '''SELECT DISTINCT p.contractor_company FROM projects p
            LEFT JOIN areas a ON a.id = p.area_id
            WHERE p.active = 1 AND p.contractor_company IS NOT NULL AND p.contractor_company != '' '''
        pp = []
        if project_id:
            pq += ' AND p.id = ?'
            pp.append(project_id)
        if area_id:
            pq += ' AND p.area_id = ?'
            pp.append(area_id)
        if division_id:
            pq += ' AND a.division_id = ?'
            pp.append(division_id)
        for r in db_fetchall(conn, pq, pp):
            names.add(r['contractor_company'])
        # From employees subcontractor field
        eq = '''SELECT DISTINCT e.subcontractor FROM employees e
            LEFT JOIN projects p ON p.id = e.project_id
            LEFT JOIN areas a ON a.id = p.area_id
            WHERE e.active = 1 AND e.subcontractor IS NOT NULL AND e.subcontractor != '' '''
        ep = []
        if project_id:
            eq += ' AND e.project_id = ?'
            ep.append(project_id)
        if area_id:
            eq += ' AND p.area_id = ?'
            ep.append(area_id)
        if division_id:
            eq += ' AND a.division_id = ?'
            ep.append(division_id)
        eq, ep = apply_project_filter(conn, eq, ep, request.user, 'e')
        for r in db_fetchall(conn, eq, ep):
            names.add(r['subcontractor'])
    finally:
        conn.close()
    return jsonify(sorted(names))


@app.route('/api/employees')
@require_auth
def api_employees():
    conn = get_db()
    try:
        search = request.args.get('search', '')
        project_id = request.args.get('project_id', '')
        area_id = request.args.get('area_id', type=int)
        division_id = request.args.get('division_id', type=int)
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 500))
        offset = (page - 1) * per_page

        query = '''SELECT e.*, p.name as project_name, a.name as area_name, d.name as division_name
            FROM employees e
            LEFT JOIN projects p ON p.id = e.project_id
            LEFT JOIN areas a ON a.id = p.area_id
            LEFT JOIN divisions d ON d.id = a.division_id
            WHERE e.active = 1'''
        params = []
        if project_id:
            query += ' AND e.project_id = ?'
            params.append(project_id)
        if area_id:
            query += ' AND p.area_id = ?'
            params.append(area_id)
        if division_id:
            query += ' AND a.division_id = ?'
            params.append(division_id)
        subcontractor = request.args.get('subcontractor', '').strip()
        if subcontractor:
            query += ' AND (e.subcontractor = ? OR e.contractor = ? OR p.contractor_company = ?)'
            params.extend([subcontractor, subcontractor, subcontractor])
        query, params = apply_project_filter(conn, query, params, request.user, 'e')
        if search:
            escaped = search.replace('%', r'\%').replace('_', r'\_')
            query += " AND (e.name LIKE ? ESCAPE '\\' OR e.employee_no LIKE ? ESCAPE '\\')"
            params.extend([f'%{escaped}%', f'%{escaped}%'])

        count_query = query.replace('SELECT e.*, p.name as project_name, a.name as area_name, d.name as division_name', 'SELECT COUNT(*) as total_count', 1)
        total_row = db_fetchone(conn, count_query, params)
        total = total_row.get('total_count', 0) if total_row else 0

        query += ' ORDER BY e.name LIMIT ? OFFSET ?'
        params.extend([per_page, offset])
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    return jsonify(rows)


@app.route('/api/scan', methods=['POST'])
@require_auth
def api_scan():
    data = request.json
    employee_no = data.get('employee_no', '').strip()
    project_id = data.get('project_id')
    site_id = data.get('site_id')
    session = data.get('session', '').upper()
    qr_data = (data.get('qr_data') or '').strip()
    user = request.user

    # Accept unique QR payload: "project_id|employee_no"
    if qr_data and '|' in qr_data:
        parts = qr_data.split('|', 1)
        try:
            project_id = int(parts[0])
            employee_no = parts[1].strip()
        except (ValueError, IndexError):
            pass

    VALID_SESSIONS = ('AM', 'PM', 'EV')  # 9 AM, 2 PM, 6 PM
    if not employee_no or not project_id or not site_id or session not in VALID_SESSIONS:
        return jsonify({'success': False, 'message': 'Missing fields'}), 400

    # Focal points (ADNOC Onshore): view-only, cannot scan
    if user.get('role') == 'focal_point':
        return jsonify({'success': False, 'message': 'View-only access. You cannot scan.'}), 403

    conn = get_db()
    try:
        if not check_project_access(conn, user, project_id):
            return jsonify({'success': False, 'message': 'No access to this area'}), 403

        site = db_fetchone(conn, 'SELECT id FROM sites WHERE id = ? AND project_id = ?', (site_id, project_id))
        if not site:
            return jsonify({'success': False, 'message': 'Invalid site for this project'}), 400

        emp = db_fetchone(conn, 'SELECT * FROM employees WHERE employee_no = ? AND project_id = ?', (employee_no, project_id))
        if not emp:
            return jsonify({'success': False, 'message': f'Not found: {employee_no}'}), 404

        today = date.today().isoformat()
        existing = db_fetchone(conn, 'SELECT id FROM attendance WHERE employee_no = ? AND project_id = ? AND scan_date = ? AND session = ?',
                               (employee_no, project_id, today, session))
        if existing:
            return jsonify({'success': False, 'duplicate': True, 'message': f'{emp["name"]} already scanned for {session}', 'employee': emp})

        # Record scanner identity (name, email, designation) on every scan
        scanner_email = user.get('email') or ''
        scanner_designation = user.get('designation') or ''
        db_execute(conn, '''INSERT INTO attendance (employee_id, employee_no, project_id, site_id, scan_date, session, supervisor_id, supervisor_name, scanner_email, scanner_designation)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (emp['id'], employee_no, project_id, site_id, today, session, user['id'], user['display_name'], scanner_email, scanner_designation))
        conn.commit()

        count = db_fetchone(conn, 'SELECT COUNT(*) as c FROM attendance WHERE project_id = ? AND site_id = ? AND scan_date = ? AND session = ?',
                            (project_id, site_id, today, session))['c']
        total = db_fetchone(conn, 'SELECT COUNT(*) as c FROM employees WHERE project_id = ? AND active = 1', (project_id,))['c']
    finally:
        conn.close()
    return jsonify({'success': True, 'message': f'Checked in: {emp["name"]}', 'employee': emp, 'site_count': count, 'site_total': total})


@app.route('/api/headcount')
@require_auth
def api_headcount():
    scan_date = request.args.get('date', date.today().isoformat())
    project_id = request.args.get('project_id')
    site_id = request.args.get('site_id')

    conn = get_db()
    try:
        query = '''SELECT p.name as project_name, s.name as site_name, a.session, COUNT(DISTINCT a.employee_no) as count
            FROM attendance a JOIN sites s ON s.id = a.site_id JOIN projects p ON p.id = a.project_id
            WHERE a.scan_date = ?'''
        params = [scan_date]
        if project_id:
            query += ' AND a.project_id = ?'
            params.append(project_id)
        if site_id:
            query += ' AND a.site_id = ?'
            params.append(site_id)
        query, params = apply_project_filter(conn, query, params, request.user)
        query += ' GROUP BY p.name, s.name, a.session ORDER BY p.name, s.name'
        rows = db_fetchall(conn, query, params)

        results = {}
        for row in rows:
            key = f"{row['project_name']}|{row['site_name']}"
            if key not in results:
                ec = db_fetchone(conn, '''SELECT COUNT(*) as c FROM employees e JOIN projects p ON p.id = e.project_id
                    WHERE e.active = 1 AND p.name = ? AND e.work_location = ?''', (row['project_name'], row['site_name']))
                results[key] = {'project': row['project_name'], 'site': row['site_name'], 'total_employees': ec['c'], 'AM': 0, 'PM': 0, 'EV': 0}
            results[key][row['session']] = row['count']

        tq = 'SELECT COUNT(*) as c FROM employees e WHERE e.active = 1'
        tp = []
        if project_id:
            tq += ' AND e.project_id = ?'
            tp.append(project_id)
        tq, tp = apply_project_filter(conn, tq, tp, request.user, 'e')
        total = db_fetchone(conn, tq, tp)['c']
    finally:
        conn.close()
    return jsonify({'date': scan_date, 'total_employees': total, 'sites': list(results.values())})


@app.route('/api/headcount/detail')
@require_auth
def api_headcount_detail():
    scan_date = request.args.get('date', date.today().isoformat())
    project_id = request.args.get('project_id')
    session = request.args.get('session', '')

    conn = get_db()
    try:
        query = '''SELECT e.name, e.employee_no, e.designation, e.discipline, a.session, a.scanned_at, a.supervisor_name, a.scanner_email, a.scanner_designation
            FROM attendance a JOIN employees e ON e.id = a.employee_id WHERE a.scan_date = ?'''
        params = [scan_date]
        if project_id:
            query += ' AND a.project_id = ?'
            params.append(project_id)
        if session:
            query += ' AND a.session = ?'
            params.append(session.upper())
        query, params = apply_project_filter(conn, query, params, request.user)
        query += ' ORDER BY a.scanned_at DESC'
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    return jsonify(rows)


@app.route('/api/headcount/missing')
@require_auth
def api_missing():
    scan_date = request.args.get('date', date.today().isoformat())
    project_id = request.args.get('project_id')
    session = request.args.get('session', 'AM')

    conn = get_db()
    try:
        query = 'SELECT e.* FROM employees e WHERE e.active = 1'
        params = []
        if project_id:
            query += ' AND e.project_id = ?'
            params.append(project_id)
        query, params = apply_project_filter(conn, query, params, request.user, 'e')
        query += ' AND e.employee_no NOT IN (SELECT a.employee_no FROM attendance a WHERE a.scan_date = ? AND a.session = ?'
        params.extend([scan_date, session.upper()])
        if project_id:
            query += ' AND a.project_id = ?'
            params.append(project_id)
        query += ') ORDER BY e.name'
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    return jsonify(rows)


@app.route('/api/stats')
@require_auth
def api_stats():
    project_id = request.args.get('project_id')
    area_id = request.args.get('area_id', type=int)
    division_id = request.args.get('division_id', type=int)
    target_date = request.args.get('date', date.today().isoformat())
    conn = get_db()
    try:
        eq = '''SELECT COUNT(*) as c FROM employees e
            LEFT JOIN projects p ON p.id = e.project_id
            LEFT JOIN areas ar ON ar.id = p.area_id
            WHERE e.active = 1'''
        ep = []
        if project_id:
            eq += ' AND e.project_id = ?'
            ep.append(project_id)
        if area_id:
            eq += ' AND p.area_id = ?'
            ep.append(area_id)
        if division_id:
            eq += ' AND ar.division_id = ?'
            ep.append(division_id)
        eq, ep = apply_project_filter(conn, eq, ep, request.user, 'e')

        def count_session(session_name):
            q = '''SELECT COUNT(DISTINCT a.employee_no) as c FROM attendance a
                LEFT JOIN projects p ON p.id = a.project_id
                LEFT JOIN areas ar ON ar.id = p.area_id
                WHERE a.scan_date = ? AND a.session = ?'''
            params = [target_date, session_name]
            if project_id:
                q += ' AND a.project_id = ?'
                params.append(project_id)
            if area_id:
                q += ' AND p.area_id = ?'
                params.append(area_id)
            if division_id:
                q += ' AND ar.division_id = ?'
                params.append(division_id)
            q, params = apply_project_filter(conn, q, params, request.user)
            return db_fetchone(conn, q, params)['c']

        total_emp = db_fetchone(conn, eq, ep)['c']
        today_am = count_session('AM')
        today_ev = count_session('EV')
        total_projects = db_fetchone(conn, 'SELECT COUNT(*) as c FROM projects WHERE active = 1')['c']
    finally:
        conn.close()
    return jsonify({'total_employees': total_emp, 'today_am': today_am, 'today_ev': today_ev, 'total_projects': total_projects, 'date': target_date})


@app.route('/api/trends')
@require_auth
def api_trends():
    """Attendance trend over N days, optionally filtered by session, designation, nationality, subcontractor."""
    days = int(request.args.get('days', 30))
    session = request.args.get('session', '').upper()
    designation = request.args.get('designation', '').strip()
    nationality = request.args.get('nationality', '').strip()
    subcontractor = request.args.get('subcontractor', '').strip()
    project_id = request.args.get('project_id', '')
    area_id = request.args.get('area_id', type=int)
    division_id = request.args.get('division_id', type=int)

    end = date.today()
    start = end - timedelta(days=days - 1)
    conn = get_db()
    try:
        query = '''SELECT att.scan_date, COUNT(DISTINCT att.employee_no) as count
            FROM attendance att
            JOIN employees e ON e.employee_no = att.employee_no AND e.project_id = att.project_id
            LEFT JOIN projects p ON p.id = att.project_id
            LEFT JOIN areas ar ON ar.id = p.area_id
            WHERE att.scan_date >= ? AND att.scan_date <= ? AND e.active = 1'''
        params = [start.isoformat(), end.isoformat()]
        if session:
            query += ' AND att.session = ?'
            params.append(session)
        if designation:
            query += ' AND e.designation = ?'
            params.append(designation)
        if nationality:
            query += ' AND e.nationality = ?'
            params.append(nationality)
        if subcontractor:
            query += ' AND (e.subcontractor = ? OR e.contractor = ? OR p.contractor_company = ?)'
            params.extend([subcontractor, subcontractor, subcontractor])
        if project_id:
            query += ' AND att.project_id = ?'
            params.append(project_id)
        if area_id:
            query += ' AND p.area_id = ?'
            params.append(area_id)
        if division_id:
            query += ' AND ar.division_id = ?'
            params.append(division_id)
        query, params = apply_project_filter(conn, query, params, request.user, 'att')
        query += ' GROUP BY att.scan_date ORDER BY att.scan_date'
        rows = db_fetchall(conn, query, params)

        # Total workforce (filtered same way)
        tq = '''SELECT COUNT(DISTINCT e.id) as c FROM employees e
            LEFT JOIN projects p ON p.id = e.project_id
            LEFT JOIN areas ar ON ar.id = p.area_id WHERE e.active = 1'''
        tp = []
        if designation:
            tq += ' AND e.designation = ?'
            tp.append(designation)
        if nationality:
            tq += ' AND e.nationality = ?'
            tp.append(nationality)
        if subcontractor:
            tq += ' AND (e.subcontractor = ? OR e.contractor = ? OR p.contractor_company = ?)'
            tp.extend([subcontractor, subcontractor, subcontractor])
        if project_id:
            tq += ' AND e.project_id = ?'
            tp.append(project_id)
        if area_id:
            tq += ' AND p.area_id = ?'
            tp.append(area_id)
        if division_id:
            tq += ' AND ar.division_id = ?'
            tp.append(division_id)
        tq, tp = apply_project_filter(conn, tq, tp, request.user, 'e')
        total = db_fetchone(conn, tq, tp)['c']

        # Distinct values for filter dropdowns
        desig_rows = db_fetchall(conn, "SELECT DISTINCT designation FROM employees WHERE active = 1 AND designation IS NOT NULL AND designation != '' ORDER BY designation")
        nat_rows = db_fetchall(conn, "SELECT DISTINCT nationality FROM employees WHERE active = 1 AND nationality IS NOT NULL AND nationality != '' ORDER BY nationality")
    finally:
        conn.close()

    dates_map = {r['scan_date']: r['count'] for r in rows}
    labels = []
    values = []
    d = start
    while d <= end:
        iso = d.isoformat()
        labels.append(iso)
        values.append(dates_map.get(iso, 0))
        d += timedelta(days=1)

    return jsonify({
        'labels': labels, 'values': values, 'total': total,
        'designations': [r['designation'] for r in desig_rows],
        'nationalities': [r['nationality'] for r in nat_rows]
    })


@app.route('/api/health-trends')
@require_auth
def api_health_trends():
    """Health metrics aggregated from employee welfare fields."""
    area_id = request.args.get('area_id', type=int)
    division_id = request.args.get('division_id', type=int)
    project_id = request.args.get('project_id', type=int)
    subcontractor = request.args.get('subcontractor', '').strip()

    conn = get_db()
    try:
        where = 'WHERE e.active = 1'
        params = []
        if project_id:
            where += ' AND e.project_id = ?'
            params.append(project_id)
        elif area_id:
            where += ' AND p.area_id = ?'
            params.append(area_id)
        elif division_id:
            where += ' AND ar.division_id = ?'
            params.append(division_id)
        if subcontractor:
            # Dropdown mixes project contractor_company and employee subcontractor — match all
            where += ' AND (e.subcontractor = ? OR e.contractor = ? OR p.contractor_company = ?)'
            params.extend([subcontractor, subcontractor, subcontractor])

        base = f'''FROM employees e
            LEFT JOIN projects p ON p.id = e.project_id
            LEFT JOIN areas ar ON ar.id = p.area_id
            {where}'''
        base, params = apply_project_filter(conn, base, list(params), request.user, 'e')

        total = db_fetchone(conn, f'SELECT COUNT(*) as c {base}', params)['c']

        # Medical result breakdown
        fit = db_fetchone(conn, f"SELECT COUNT(*) as c {base} AND LOWER(COALESCE(e.medical_result,'')) LIKE '%fit%' AND LOWER(COALESCE(e.medical_result,'')) NOT LIKE '%unfit%'", params)['c']
        unfit = db_fetchone(conn, f"SELECT COUNT(*) as c {base} AND LOWER(COALESCE(e.medical_result,'')) LIKE '%unfit%'", params)['c']
        no_result = max(0, total - fit - unfit)

        # Medical overdue (next_medical_due < today)
        today_str = date.today().isoformat()
        overdue = db_fetchone(conn, f"SELECT COUNT(*) as c {base} AND e.next_medical_due IS NOT NULL AND TRIM(COALESCE(e.next_medical_due,'')) != '' AND e.next_medical_due < ?", params + [today_str])['c']

        # Chronic conditions
        has_chronic = db_fetchone(conn, f"SELECT COUNT(*) as c {base} AND e.chronic_condition IS NOT NULL AND TRIM(e.chronic_condition) != '' AND LOWER(e.chronic_condition) NOT LIKE '%nil%' AND LOWER(e.chronic_condition) NOT LIKE '%none%' AND LOWER(e.chronic_condition) NOT LIKE '%no%' AND LOWER(e.chronic_condition) != 'n/a'", params)['c']
        chronic_treated = db_fetchone(conn, f"SELECT COUNT(*) as c {base} AND e.chronic_treated IS NOT NULL AND (LOWER(e.chronic_treated) LIKE '%yes%' OR LOWER(e.chronic_treated) LIKE '%control%')", params)['c']
        chronic_untreated = has_chronic - chronic_treated if has_chronic > chronic_treated else 0

        # General feeling
        feel_good = db_fetchone(conn, f"SELECT COUNT(*) as c {base} AND e.general_feeling IS NOT NULL AND (LOWER(e.general_feeling) LIKE '%good%' OR LOWER(e.general_feeling) LIKE '%fine%' OR LOWER(e.general_feeling) LIKE '%well%' OR LOWER(e.general_feeling) LIKE '%excellent%')", params)['c']
        feel_bad = db_fetchone(conn, f"SELECT COUNT(*) as c {base} AND e.general_feeling IS NOT NULL AND (LOWER(e.general_feeling) LIKE '%bad%' OR LOWER(e.general_feeling) LIKE '%poor%' OR LOWER(e.general_feeling) LIKE '%sick%' OR LOWER(e.general_feeling) LIKE '%unwell%' OR LOWER(e.general_feeling) LIKE '%tired%')", params)['c']
        feel_neutral = max(0, total - feel_good - feel_bad)

        # Medical frequency breakdown
        freq_rows = db_fetchall(conn, f"SELECT e.medical_frequency, COUNT(*) as c {base} AND e.medical_frequency IS NOT NULL AND TRIM(COALESCE(e.medical_frequency,'')) != '' GROUP BY e.medical_frequency ORDER BY c DESC", params)

        # Chronic condition types
        chronic_rows = db_fetchall(conn, f"SELECT e.chronic_condition, COUNT(*) as c {base} AND e.chronic_condition IS NOT NULL AND TRIM(e.chronic_condition) != '' AND LOWER(e.chronic_condition) NOT LIKE '%nil%' AND LOWER(e.chronic_condition) NOT LIKE '%none%' AND LOWER(e.chronic_condition) NOT LIKE '%no%' AND LOWER(e.chronic_condition) != 'n/a' GROUP BY e.chronic_condition ORDER BY c DESC LIMIT 15", params)

        # Per-person risk list (those with any flag)
        risk_sql = f'''SELECT e.name, e.employee_no, e.designation, e.nationality,
            e.medical_result, e.next_medical_due, e.chronic_condition, e.chronic_treated,
            e.general_feeling, e.medical_frequency, p.name as project_name
            {base}
            AND (
                LOWER(COALESCE(e.medical_result,'')) LIKE '%unfit%'
                OR (e.next_medical_due IS NOT NULL AND TRIM(COALESCE(e.next_medical_due,'')) != '' AND e.next_medical_due < ?)
                OR (e.chronic_condition IS NOT NULL AND TRIM(e.chronic_condition) != '' AND LOWER(e.chronic_condition) NOT LIKE '%nil%' AND LOWER(e.chronic_condition) NOT LIKE '%none%' AND LOWER(e.chronic_condition) NOT LIKE '%no%' AND LOWER(e.chronic_condition) != 'n/a')
                OR (e.general_feeling IS NOT NULL AND (LOWER(e.general_feeling) LIKE '%bad%' OR LOWER(e.general_feeling) LIKE '%poor%' OR LOWER(e.general_feeling) LIKE '%sick%' OR LOWER(e.general_feeling) LIKE '%unwell%'))
            )
            ORDER BY e.name LIMIT 500'''
        risk_people = db_fetchall(conn, risk_sql, params + [today_str])
    finally:
        conn.close()

    return jsonify({
        'total': total,
        'medical': {'fit': fit, 'unfit': unfit, 'no_result': no_result, 'overdue': overdue},
        'chronic': {'has_chronic': has_chronic, 'treated': chronic_treated, 'untreated': chronic_untreated},
        'feeling': {'good': feel_good, 'bad': feel_bad, 'neutral': feel_neutral},
        'medical_freq': [{'label': r['medical_frequency'], 'count': r['c']} for r in freq_rows],
        'chronic_types': [{'label': r['chronic_condition'], 'count': r['c']} for r in chronic_rows],
        'risk_people': [dict(r) for r in risk_people],
    })


# ============================================================
# TWL (Thermal Work Limit)
# ============================================================

def compute_twl_risk_zone(twl_value):
    if twl_value >= 140:
        return 'low'
    elif twl_value >= 115:
        return 'medium'
    else:
        return 'high'

TWL_ZONES = {
    'low': {
        'label': 'Low Risk — Unrestricted',
        'twl_range': '140 – 220',
        'color': '#22c55e',
        'interventions': 'No limits on self-paced work for educated, hydrated workers',
        'rehydration': {'light': '600 mL – 1 Liter/hour'},
        'work_rest': {'light': 'Safe for all continuous self-paced work'},
    },
    'medium': {
        'label': 'Medium Risk — Cautionary',
        'twl_range': '115 – 140',
        'color': '#f59e0b',
        'interventions': 'Environmental conditions require additional precautions; implement practicable engineering control measures to reduce heat stress (e.g. shade, ventilation)',
        'rehydration': {'light': '1 – 1.2 Liters/hour', 'heavy': '> 1.2 Liters/hour'},
        'work_rest': {'light': 'Safe for continuous self-paced light work', 'heavy': '45 min work – 15 min rest'},
    },
    'high': {
        'label': 'High Risk',
        'twl_range': '< 115',
        'color': '#ef4444',
        'interventions': 'Strict work/rest cycling required; No person to work alone; No unacclimatized person to work; Induction required; Provide personal water bottle (2L) on-site',
        'rehydration': {'all': '> 1.2 Liters/hour'},
        'work_rest': {'light': '45 min work – 15 min rest', 'heavy': '20 min work – 40 min rest'},
    },
}


@app.route('/api/twl', methods=['POST'])
@require_auth
def api_twl_record():
    data = request.get_json(force=True)
    twl_value = data.get('twl_value')
    area_id = data.get('area_id')
    site_id = data.get('site_id')
    reading_date = data.get('reading_date', date.today().isoformat())
    reading_time = data.get('reading_time', datetime.now().strftime('%H:%M'))
    temperature = data.get('temperature')
    humidity = data.get('humidity')
    wind_speed = data.get('wind_speed')
    work_type = data.get('work_type', 'light')
    notes = data.get('notes', '')

    if twl_value is None:
        return jsonify({'success': False, 'message': 'TWL value is required'}), 400
    try:
        twl_value = float(twl_value)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'TWL value must be a number'}), 400

    risk_zone = compute_twl_risk_zone(twl_value)

    conn = get_db()
    try:
        db_execute(conn, '''INSERT INTO twl_readings
            (area_id, site_id, reading_date, reading_time, twl_value, risk_zone,
             temperature, humidity, wind_speed, work_type, recorded_by, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (area_id, site_id, reading_date, reading_time, twl_value, risk_zone,
             temperature, humidity, wind_speed, work_type, request.user['id'], notes))
        conn.commit()
    finally:
        conn.close()

    zone_info = TWL_ZONES[risk_zone]
    audit(request.user['id'], 'twl_reading', f'TWL={twl_value} zone={risk_zone} area={area_id}')
    return jsonify({
        'success': True,
        'message': f'TWL reading recorded: {twl_value} ({zone_info["label"]})',
        'risk_zone': risk_zone,
        'zone_info': zone_info
    })


@app.route('/api/twl')
@require_auth
def api_twl_list():
    area_id = request.args.get('area_id', type=int)
    division_id = request.args.get('division_id', type=int)
    days = int(request.args.get('days', 30))
    end = date.today()
    start = end - timedelta(days=days - 1)

    conn = get_db()
    try:
        query = '''SELECT t.*, a.name as area_name, u.display_name as recorded_by_name
            FROM twl_readings t
            LEFT JOIN areas a ON a.id = t.area_id
            LEFT JOIN users u ON u.id = t.recorded_by
            WHERE t.reading_date >= ? AND t.reading_date <= ?'''
        params = [start.isoformat(), end.isoformat()]
        if area_id:
            query += ' AND t.area_id = ?'
            params.append(area_id)
        elif division_id:
            query += ' AND a.division_id = ?'
            params.append(division_id)
        query += ' ORDER BY t.reading_date DESC, t.reading_time DESC'
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/twl/summary')
@require_auth
def api_twl_summary():
    area_id = request.args.get('area_id', type=int)
    division_id = request.args.get('division_id', type=int)
    days = int(request.args.get('days', 30))
    end = date.today()
    start = end - timedelta(days=days - 1)
    today_str = date.today().isoformat()

    conn = get_db()
    try:
        where = 'WHERE t.reading_date >= ? AND t.reading_date <= ?'
        params = [start.isoformat(), end.isoformat()]
        if area_id:
            where += ' AND t.area_id = ?'
            params.append(area_id)
        elif division_id:
            where += ' AND a.division_id = ?'
            params.append(division_id)

        base = f'''FROM twl_readings t
            LEFT JOIN areas a ON a.id = t.area_id
            {where}'''

        total_readings = db_fetchone(conn, f'SELECT COUNT(*) as c {base}', params)['c']

        zone_counts = db_fetchall(conn, f'SELECT risk_zone, COUNT(*) as c {base} GROUP BY risk_zone', params)
        zones = {r['risk_zone']: r['c'] for r in zone_counts}

        avg_twl = db_fetchone(conn, f'SELECT AVG(twl_value) as avg_val, MIN(twl_value) as min_val, MAX(twl_value) as max_val {base}', params)

        today_params = [today_str, today_str]
        if area_id:
            today_params.append(area_id)
        elif division_id:
            today_params.append(division_id)
        today_where = where.replace(start.isoformat(), today_str).replace(end.isoformat(), today_str)
        today_base = f'''FROM twl_readings t LEFT JOIN areas a ON a.id = t.area_id {today_where}'''
        today_readings = db_fetchall(conn, f'SELECT t.twl_value, t.risk_zone, t.reading_time, a.name as area_name {today_base} ORDER BY t.reading_time DESC', today_params)

        trend = db_fetchall(conn, f'SELECT t.reading_date, AVG(t.twl_value) as avg_twl, MIN(t.twl_value) as min_twl, MAX(t.twl_value) as max_twl, COUNT(*) as readings {base} GROUP BY t.reading_date ORDER BY t.reading_date', params)

        high_risk_days = db_fetchone(conn, f"SELECT COUNT(DISTINCT t.reading_date) as c {base} AND t.risk_zone = 'high'", params)['c']
    finally:
        conn.close()

    return jsonify({
        'total_readings': total_readings,
        'zones': zones,
        'avg_twl': round(avg_twl['avg_val'] or 0, 1),
        'min_twl': round(avg_twl['min_val'] or 0, 1),
        'max_twl': round(avg_twl['max_val'] or 0, 1),
        'today': [dict(r) for r in today_readings],
        'trend': [dict(r) for r in trend],
        'high_risk_days': high_risk_days,
        'zone_definitions': TWL_ZONES,
    })


# ============================================================
# Anomaly Detection Engine
# ============================================================

@app.route('/api/anomalies')
@require_auth
def api_anomalies():
    """Detect anomalies across attendance, health, O&I, TWL, and cross-domain patterns."""
    target_date = request.args.get('date', date.today().isoformat())
    division_id = request.args.get('division_id', type=int)
    area_id = request.args.get('area_id', type=int)
    project_id = request.args.get('project_id', type=int)

    conn = get_db()
    anomalies = []
    try:
        # Scope filters
        emp_where = 'WHERE e.active = 1'
        emp_params = []
        if project_id:
            emp_where += ' AND e.project_id = ?'
            emp_params.append(project_id)
        elif area_id:
            emp_where += ' AND p.area_id = ?'
            emp_params.append(area_id)
        elif division_id:
            emp_where += ' AND ar.division_id = ?'
            emp_params.append(division_id)

        emp_base = f'''FROM employees e
            LEFT JOIN projects p ON p.id = e.project_id
            LEFT JOIN areas ar ON ar.id = p.area_id
            {emp_where}'''

        total_emp = db_fetchone(conn, f'SELECT COUNT(*) as c {emp_base}', emp_params)['c']

        # ── 1. ATTENDANCE ANOMALIES ──

        # Today's attendance
        for sess_name, sess_label in [('AM', '9 AM'), ('EV', '7 PM')]:
            att_q = f'''SELECT COUNT(DISTINCT att.employee_no) as c
                FROM attendance att
                JOIN employees e ON e.employee_no = att.employee_no AND e.project_id = att.project_id
                LEFT JOIN projects p ON p.id = att.project_id
                LEFT JOIN areas ar ON ar.id = p.area_id
                WHERE att.scan_date = ? AND att.session = ? AND e.active = 1'''
            att_p = [target_date, sess_name]
            if project_id:
                att_q += ' AND att.project_id = ?'
                att_p.append(project_id)
            elif area_id:
                att_q += ' AND p.area_id = ?'
                att_p.append(area_id)
            elif division_id:
                att_q += ' AND ar.division_id = ?'
                att_p.append(division_id)
            today_count = db_fetchone(conn, att_q, att_p)['c']

            # 7-day rolling average
            d7 = (date.fromisoformat(target_date) - timedelta(days=7)).isoformat()
            avg_q = att_q.replace('att.scan_date = ?', 'att.scan_date >= ? AND att.scan_date < ?')
            avg_p = [d7, target_date, sess_name] + att_p[2:]
            avg_row = db_fetchone(conn, avg_q.replace('COUNT(DISTINCT att.employee_no) as c', 'COUNT(DISTINCT att.employee_no || att.scan_date) as c'), avg_p)
            avg_7d = (avg_row['c'] / 7) if avg_row else 0

            if total_emp > 0:
                rate = (today_count / total_emp) * 100
                if rate < 50 and today_count > 0:
                    anomalies.append({
                        'category': 'attendance', 'severity': 'high',
                        'title': f'Low {sess_label} attendance',
                        'detail': f'Only {today_count}/{total_emp} ({rate:.0f}%) scanned at {sess_label}',
                        'metric': f'{rate:.0f}%'
                    })
                elif rate < 75 and today_count > 0:
                    anomalies.append({
                        'category': 'attendance', 'severity': 'medium',
                        'title': f'Below-average {sess_label} attendance',
                        'detail': f'{today_count}/{total_emp} ({rate:.0f}%) scanned — below 75% threshold',
                        'metric': f'{rate:.0f}%'
                    })

            if avg_7d > 0 and today_count > 0 and today_count < avg_7d * 0.75:
                drop_pct = ((avg_7d - today_count) / avg_7d) * 100
                anomalies.append({
                    'category': 'attendance', 'severity': 'high',
                    'title': f'{sess_label} attendance drop vs 7-day average',
                    'detail': f'{today_count} present vs {avg_7d:.0f} avg ({drop_pct:.0f}% drop)',
                    'metric': f'-{drop_pct:.0f}%'
                })

        # AM vs PM gap
        am_q = f'''SELECT COUNT(DISTINCT att.employee_no) as c
            FROM attendance att JOIN employees e ON e.employee_no = att.employee_no AND e.project_id = att.project_id
            LEFT JOIN projects p ON p.id = att.project_id LEFT JOIN areas ar ON ar.id = p.area_id
            WHERE att.scan_date = ? AND att.session = 'AM' AND e.active = 1'''
        ev_q = am_q.replace("'AM'", "'EV'")
        am_p = [target_date]
        if project_id:
            am_q += ' AND att.project_id = ?'
            ev_q += ' AND att.project_id = ?'
            am_p.append(project_id)
        elif area_id:
            am_q += ' AND p.area_id = ?'
            ev_q += ' AND p.area_id = ?'
            am_p.append(area_id)
        elif division_id:
            am_q += ' AND ar.division_id = ?'
            ev_q += ' AND ar.division_id = ?'
            am_p.append(division_id)
        am_count = db_fetchone(conn, am_q, am_p)['c']
        ev_count = db_fetchone(conn, ev_q, am_p)['c']
        if am_count > 10 and ev_count > 0 and ev_count < am_count * 0.7:
            gap_pct = ((am_count - ev_count) / am_count) * 100
            anomalies.append({
                'category': 'attendance', 'severity': 'medium',
                'title': 'AM-to-PM attendance drop',
                'detail': f'{am_count} present at 9AM but only {ev_count} at 7PM ({gap_pct:.0f}% left early)',
                'metric': f'-{gap_pct:.0f}%'
            })

        # Per-project low attendance
        proj_att = db_fetchall(conn, f'''SELECT p.name as pname, COUNT(DISTINCT att.employee_no) as present,
            (SELECT COUNT(*) FROM employees e2 WHERE e2.project_id = p.id AND e2.active = 1) as total
            FROM attendance att
            JOIN projects p ON p.id = att.project_id
            WHERE att.scan_date = ? AND att.session = 'AM'
            GROUP BY p.id HAVING total > 5 AND present < total * 0.4
            ORDER BY (CAST(present AS REAL) / total) ASC LIMIT 5''', [target_date])
        for pa in proj_att:
            pct = (pa['present'] / pa['total'] * 100) if pa['total'] else 0
            anomalies.append({
                'category': 'attendance', 'severity': 'high',
                'title': f'Critical attendance: {pa["pname"]}',
                'detail': f'Only {pa["present"]}/{pa["total"]} ({pct:.0f}%) present at 9 AM',
                'metric': f'{pct:.0f}%'
            })

        # ── 2. HEALTH ANOMALIES ──

        unfit = db_fetchone(conn, f"SELECT COUNT(*) as c {emp_base} AND LOWER(e.medical_result) LIKE '%unfit%'", emp_params)['c']
        if unfit > 0:
            anomalies.append({
                'category': 'health', 'severity': 'high',
                'title': f'{unfit} worker{"s" if unfit > 1 else ""} marked UNFIT',
                'detail': 'Medically unfit workers still in active roster — immediate review required',
                'metric': str(unfit)
            })

        overdue = db_fetchone(conn, f"SELECT COUNT(*) as c {emp_base} AND e.next_medical_due != '' AND e.next_medical_due IS NOT NULL AND e.next_medical_due < ?", emp_params + [target_date])['c']
        if overdue > 10:
            anomalies.append({
                'category': 'health', 'severity': 'high',
                'title': f'{overdue} overdue medical exams',
                'detail': 'Workers with expired medical clearance — schedule exams urgently',
                'metric': str(overdue)
            })
        elif overdue > 0:
            anomalies.append({
                'category': 'health', 'severity': 'medium',
                'title': f'{overdue} overdue medical exam{"s" if overdue > 1 else ""}',
                'detail': 'Workers due for medical re-examination',
                'metric': str(overdue)
            })

        chronic = db_fetchone(conn, f"SELECT COUNT(*) as c {emp_base} AND e.chronic_condition IS NOT NULL AND e.chronic_condition != '' AND LOWER(e.chronic_condition) NOT IN ('nil','none','no','n/a','')", emp_params)['c']
        chronic_untreated = db_fetchone(conn, f"SELECT COUNT(*) as c {emp_base} AND e.chronic_condition IS NOT NULL AND e.chronic_condition != '' AND LOWER(e.chronic_condition) NOT IN ('nil','none','no','n/a','') AND (e.chronic_treated IS NULL OR e.chronic_treated = '' OR LOWER(e.chronic_treated) NOT LIKE '%yes%')", emp_params)['c']
        if chronic_untreated > 5:
            anomalies.append({
                'category': 'health', 'severity': 'high',
                'title': f'{chronic_untreated} untreated chronic conditions',
                'detail': f'Out of {chronic} workers with chronic conditions, {chronic_untreated} are not confirmed under treatment',
                'metric': str(chronic_untreated)
            })

        # ── 3. O&I (SAFETY) ANOMALIES ──

        d7_start = (date.fromisoformat(target_date) - timedelta(days=7)).isoformat()
        obs_where = "WHERE o.observation_date = ?"
        obs_params_day = [target_date]
        obs_where_7d = "WHERE o.observation_date >= ? AND o.observation_date <= ?"
        obs_params_7d = [d7_start, target_date]
        if project_id:
            obs_where += ' AND o.project_id = ?'
            obs_where_7d += ' AND o.project_id = ?'
            obs_params_day.append(project_id)
            obs_params_7d.append(project_id)
        elif area_id:
            obs_where += ' AND o.area_id = ?'
            obs_where_7d += ' AND o.area_id = ?'
            obs_params_day.append(area_id)
            obs_params_7d.append(area_id)
        elif division_id:
            obs_where += ' AND o.division_id = ?'
            obs_where_7d += ' AND o.division_id = ?'
            obs_params_day.append(division_id)
            obs_params_7d.append(division_id)

        # HIPO events today
        hipo = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o {obs_where} AND o.observation_group = 'HIPO'", obs_params_day)['c']
        if hipo > 0:
            anomalies.append({
                'category': 'safety', 'severity': 'critical',
                'title': f'{hipo} HIPO event{"s" if hipo > 1 else ""} recorded',
                'detail': 'High Potential Incident/Observation — requires immediate management attention',
                'metric': str(hipo)
            })

        # Near misses today
        near_miss = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o {obs_where} AND o.observation_group = 'Near Miss'", obs_params_day)['c']
        if near_miss >= 3:
            anomalies.append({
                'category': 'safety', 'severity': 'high',
                'title': f'{near_miss} near misses today',
                'detail': 'Elevated near-miss count — investigate root causes and patterns',
                'metric': str(near_miss)
            })

        # Unsafe observation spike vs 7-day average
        unsafe_today = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o {obs_where} AND o.observation_group NOT IN ('Safe Act','Safe Condition')", obs_params_day)['c']
        unsafe_7d = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o {obs_where_7d} AND o.observation_group NOT IN ('Safe Act','Safe Condition')", obs_params_7d)['c']
        unsafe_avg = unsafe_7d / 7 if unsafe_7d else 0
        if unsafe_avg > 0 and unsafe_today > unsafe_avg * 1.5 and unsafe_today >= 3:
            anomalies.append({
                'category': 'safety', 'severity': 'high',
                'title': 'Unsafe observation spike',
                'detail': f'{unsafe_today} unsafe observations today vs {unsafe_avg:.1f} daily average — {((unsafe_today / unsafe_avg - 1) * 100):.0f}% above normal',
                'metric': f'+{((unsafe_today / unsafe_avg - 1) * 100):.0f}%'
            })

        # Repeat observation types (same type appearing 3+ times in 7 days at same project)
        repeat_types = db_fetchall(conn, f'''SELECT o.observation_type, p.name as pname, COUNT(*) as c
            FROM observations o LEFT JOIN projects p ON p.id = o.project_id
            {obs_where_7d} AND o.observation_group NOT IN ('Safe Act','Safe Condition')
            AND o.observation_type IS NOT NULL AND o.observation_type != ''
            GROUP BY o.observation_type, o.project_id HAVING c >= 3
            ORDER BY c DESC LIMIT 3''', obs_params_7d)
        for rt in repeat_types:
            anomalies.append({
                'category': 'safety', 'severity': 'medium',
                'title': f'Recurring: {rt["observation_type"]}',
                'detail': f'Observed {rt["c"]} times in 7 days{" at " + rt["pname"] if rt["pname"] else ""} — systemic issue likely',
                'metric': f'{rt["c"]}x'
            })

        # Safe-to-unsafe ratio declining
        d14_start = (date.fromisoformat(target_date) - timedelta(days=14)).isoformat()
        week1_safe = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o WHERE o.observation_date >= ? AND o.observation_date < ? AND o.observation_group IN ('Safe Act','Safe Condition')", [d14_start, d7_start])['c']
        week1_total = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o WHERE o.observation_date >= ? AND o.observation_date < ?", [d14_start, d7_start])['c']
        week2_safe = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o WHERE o.observation_date >= ? AND o.observation_date <= ? AND o.observation_group IN ('Safe Act','Safe Condition')", [d7_start, target_date])['c']
        week2_total = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o WHERE o.observation_date >= ? AND o.observation_date <= ?", [d7_start, target_date])['c']
        if week1_total > 5 and week2_total > 5:
            ratio1 = week1_safe / week1_total
            ratio2 = week2_safe / week2_total
            if ratio2 < ratio1 * 0.8:
                anomalies.append({
                    'category': 'safety', 'severity': 'medium',
                    'title': 'Safety ratio declining',
                    'detail': f'Safe observations dropped from {ratio1 * 100:.0f}% to {ratio2 * 100:.0f}% week-over-week',
                    'metric': f'{ratio2 * 100:.0f}%'
                })

        # ── 4. TWL ANOMALIES ──

        twl_high = db_fetchall(conn, "SELECT t.twl_value, t.risk_zone, a.name as area_name FROM twl_readings t LEFT JOIN areas a ON a.id = t.area_id WHERE t.reading_date = ? AND t.risk_zone = 'high' ORDER BY t.twl_value ASC LIMIT 3", [target_date])
        for tr in twl_high:
            anomalies.append({
                'category': 'twl', 'severity': 'critical',
                'title': f'High-risk TWL: {tr["twl_value"]}',
                'detail': f'TWL reading at {tr["area_name"] or "unknown area"} — work-rest cycles mandatory, restrict heavy work',
                'metric': str(tr['twl_value'])
            })

        twl_medium = db_fetchone(conn, "SELECT COUNT(*) as c FROM twl_readings WHERE reading_date = ? AND risk_zone = 'medium'", [target_date])['c']
        if twl_medium > 0 and not twl_high:
            anomalies.append({
                'category': 'twl', 'severity': 'medium',
                'title': f'{twl_medium} medium-risk TWL reading{"s" if twl_medium > 1 else ""}',
                'detail': 'Cautionary conditions — ensure hydration protocols and paced work schedules',
                'metric': str(twl_medium)
            })

        # ── 5. CROSS-DOMAIN CORRELATIONS ──

        # High TWL + high attendance = heat stress risk
        if twl_high and am_count > total_emp * 0.7 and total_emp > 20:
            anomalies.append({
                'category': 'cross', 'severity': 'critical',
                'title': 'Heat stress risk — high TWL + full attendance',
                'detail': f'{am_count} workers on site during high-risk TWL conditions — consider stand-down or shift reduction',
                'metric': '⚠'
            })

        # Unsafe spike at specific project with low attendance (possible safety culture issue)
        if unsafe_today >= 3:
            unsafe_by_proj = db_fetchall(conn, f'''SELECT p.name as pname, COUNT(*) as c
                FROM observations o LEFT JOIN projects p ON p.id = o.project_id
                {obs_where} AND o.observation_group NOT IN ('Safe Act','Safe Condition')
                GROUP BY o.project_id HAVING c >= 2 ORDER BY c DESC LIMIT 2''', obs_params_day)
            for up in unsafe_by_proj:
                if up['pname']:
                    anomalies.append({
                        'category': 'cross', 'severity': 'high',
                        'title': f'Safety hotspot: {up["pname"]}',
                        'detail': f'{up["c"]} unsafe observations concentrated at this project today',
                        'metric': str(up['c'])
                    })

    finally:
        conn.close()

    # Sort: critical first, then high, medium, low
    severity_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
    anomalies.sort(key=lambda a: severity_order.get(a['severity'], 9))

    return jsonify({
        'date': target_date,
        'count': len(anomalies),
        'anomalies': anomalies,
    })


def _linear_slope_y(values):
    """Least-squares slope for y vs x=0..n-1. Returns 0 if degenerate."""
    n = len(values)
    if n < 2:
        return 0.0
    xs = list(range(n))
    mx = sum(xs) / n
    my = sum(values) / n
    var_x = sum((x - mx) ** 2 for x in xs)
    if var_x < 1e-9:
        return 0.0
    cov = sum((xs[i] - mx) * (values[i] - my) for i in range(n))
    return cov / var_x


def _risk_level_from_index(score):
    if score >= 75:
        return 'critical'
    if score >= 50:
        return 'high'
    if score >= 25:
        return 'medium'
    return 'low'


@app.route('/api/risk-engine')
@require_auth
def api_risk_engine():
    """
    Predictive / preventive risk engine: trend analysis, composite risk index,
    leading indicators, and prioritized preventive actions (runs on live + demo data).
    """
    target_date = request.args.get('date', date.today().isoformat())
    division_id = request.args.get('division_id', type=int)
    area_id = request.args.get('area_id', type=int)
    project_id = request.args.get('project_id', type=int)

    td = date.fromisoformat(target_date)
    d14 = (td - timedelta(days=13)).isoformat()
    d7 = (td - timedelta(days=6)).isoformat()
    d_prev7_start = (td - timedelta(days=13)).isoformat()
    d_prev7_end = (td - timedelta(days=7)).isoformat()

    conn = get_db()
    predictive_signals = []
    preventive_recommendations = []
    domains = {}

    try:
        emp_where = 'WHERE e.active = 1'
        emp_params = []
        if project_id:
            emp_where += ' AND e.project_id = ?'
            emp_params.append(project_id)
        elif area_id:
            emp_where += ' AND p.area_id = ?'
            emp_params.append(area_id)
        elif division_id:
            emp_where += ' AND ar.division_id = ?'
            emp_params.append(division_id)

        emp_base = f'''FROM employees e
            LEFT JOIN projects p ON p.id = e.project_id
            LEFT JOIN areas ar ON ar.id = p.area_id
            {emp_where}'''
        total_emp = db_fetchone(conn, f'SELECT COUNT(*) as c {emp_base}', emp_params)['c'] or 1

        # --- Attendance: daily AM presence rate (14d) ---
        att_join = '''FROM attendance att
            JOIN employees e ON e.employee_no = att.employee_no AND e.project_id = att.project_id
            LEFT JOIN projects p ON p.id = att.project_id
            LEFT JOIN areas ar ON ar.id = p.area_id
            WHERE att.scan_date >= ? AND att.scan_date <= ? AND att.session = ? AND e.active = 1'''
        att_p_base = [d14, target_date, 'AM']
        if project_id:
            att_join += ' AND att.project_id = ?'
            att_p_base.append(project_id)
        elif area_id:
            att_join += ' AND p.area_id = ?'
            att_p_base.append(area_id)
        elif division_id:
            att_join += ' AND ar.division_id = ?'
            att_p_base.append(division_id)

        daily_att = db_fetchall(conn, f'SELECT att.scan_date as d, COUNT(DISTINCT att.employee_no) as c {att_join} GROUP BY att.scan_date ORDER BY att.scan_date', att_p_base)
        att_map = {row['d'] if isinstance(row, dict) else row[0]: (row['c'] if isinstance(row, dict) else row[1]) for row in daily_att}
        rates = []
        cur = date.fromisoformat(d14)
        end_d = td
        while cur <= end_d:
            ds = cur.isoformat()
            rates.append((att_map.get(ds, 0) / total_emp) * 100)
            cur += timedelta(days=1)
        att_slope = _linear_slope_y(rates) if rates else 0
        last3 = sum(rates[-3:]) / 3 if len(rates) >= 3 else (rates[-1] if rates else 0)
        first3 = sum(rates[:3]) / 3 if len(rates) >= 3 else (rates[0] if rates else 0)
        att_decline_pp = first3 - last3

        att_score = 0
        if last3 < 60 and last3 > 0:
            att_score += min(15, int((60 - last3) * 0.4))
        if att_slope < -0.35:
            att_score += min(12, int(abs(att_slope) * 8))
        if att_decline_pp > 8:
            att_score += min(10, int(att_decline_pp * 0.5))
        att_score = min(25, att_score)

        att_trend = 'worsening' if att_slope < -0.2 or att_decline_pp > 5 else ('improving' if att_slope > 0.2 else 'stable')
        domains['attendance'] = {
            'score': att_score,
            'trend': att_trend,
            'summary': f'Recent AM presence ~{last3:.0f}% of roster (14d trend slope {att_slope:+.2f} pp/day)',
        }
        if att_slope < -0.25 and len(rates) >= 7:
            predictive_signals.append({
                'type': 'predictive',
                'severity': 'medium',
                'title': 'Attendance momentum declining',
                'detail': f'14-day trend suggests falling scan-in rates (~{att_decline_pp:.1f} pp drop vs period start). Early signal of engagement or logistics issues.',
                'horizon_days': 14,
                'confidence': min(0.9, 0.45 + min(0.4, abs(att_slope) * 0.15)),
            })
            preventive_recommendations.append({
                'priority': 2,
                'domain': 'attendance',
                'action': 'Review scanner coverage, transport, and roster changes; communicate expectations for AM sign-in.',
                'rationale': 'Leading indicator: attendance trajectory negative before thresholds breach.',
            })

        # --- Health pipeline ---
        unfit = db_fetchone(conn, f"SELECT COUNT(*) as c {emp_base} AND LOWER(e.medical_result) LIKE '%unfit%'", emp_params)['c']
        overdue = db_fetchone(conn, f"SELECT COUNT(*) as c {emp_base} AND e.next_medical_due IS NOT NULL AND e.next_medical_due != '' AND e.next_medical_due < ?", emp_params + [target_date])['c']
        due_30 = db_fetchone(conn, f"SELECT COUNT(*) as c {emp_base} AND e.next_medical_due IS NOT NULL AND e.next_medical_due != '' AND e.next_medical_due > ? AND e.next_medical_due <= ?", emp_params + [target_date, (td + timedelta(days=30)).isoformat()])['c']
        chronic_ut = db_fetchone(conn, f"SELECT COUNT(*) as c {emp_base} AND e.chronic_condition IS NOT NULL AND e.chronic_condition != '' AND LOWER(e.chronic_condition) NOT IN ('nil','none','no','n/a') AND (e.chronic_treated IS NULL OR e.chronic_treated = '' OR LOWER(e.chronic_treated) NOT LIKE '%yes%')", emp_params)['c']

        health_score = min(20, unfit * 4 + min(12, overdue // 3) + min(8, chronic_ut // 2) + min(6, due_30 // 25))
        domains['health'] = {
            'score': health_score,
            'trend': 'stable',
            'summary': f'{unfit} unfit, {overdue} overdue medicals, {due_30} due within 30 days, {chronic_ut} chronic not confirmed treated',
        }
        if due_30 > 50:
            predictive_signals.append({
                'type': 'preventive',
                'severity': 'low',
                'title': 'Medical renewal wave approaching',
                'detail': f'{due_30} workers have medical due in the next 30 days — capacity planning now avoids compliance spikes.',
                'horizon_days': 30,
                'confidence': 0.85,
            })
            preventive_recommendations.append({
                'priority': 3,
                'domain': 'health',
                'action': 'Schedule batch medical slots and notify contractors of renewal windows.',
                'rationale': 'Preventive scheduling reduces overdue risk and last-minute unfit flags.',
            })
        if unfit > 0:
            preventive_recommendations.append({
                'priority': 1,
                'domain': 'health',
                'action': 'Immediate review of UNFIT workers: restrict site access until cleared.',
                'rationale': 'Active unfit roster is a direct regulatory and duty-of-care exposure.',
            })

        # --- O&I: unsafe trend ---
        obs_where = 'WHERE o.observation_date >= ? AND o.observation_date <= ?'
        obs_p = [d14, target_date]
        if project_id:
            obs_where += ' AND o.project_id = ?'
            obs_p.append(project_id)
        elif area_id:
            obs_where += ' AND o.area_id = ?'
            obs_p.append(area_id)
        elif division_id:
            obs_where += ' AND o.division_id = ?'
            obs_p.append(division_id)

        daily_unsafe = db_fetchall(conn, f'''SELECT o.observation_date as d, COUNT(*) as c
            FROM observations o {obs_where}
            AND o.observation_group NOT IN ('Safe Act','Safe Condition')
            GROUP BY o.observation_date ORDER BY o.observation_date''', obs_p)
        umap = {row['d'] if isinstance(row, dict) else row[0]: (row['c'] if isinstance(row, dict) else row[1]) for row in daily_unsafe}
        unsafe_series = []
        cur = date.fromisoformat(d14)
        while cur <= td:
            unsafe_series.append(umap.get(cur.isoformat(), 0))
            cur += timedelta(days=1)
        unsafe_slope = _linear_slope_y(unsafe_series) if unsafe_series else 0
        w2_unsafe = sum(unsafe_series[-7:]) if len(unsafe_series) >= 7 else sum(unsafe_series)
        w1_unsafe = sum(unsafe_series[-14:-7]) if len(unsafe_series) >= 14 else w2_unsafe
        unsafe_accel = (w2_unsafe - w1_unsafe) / max(1, w1_unsafe) if w1_unsafe else (1.0 if w2_unsafe > 3 else 0)

        hipo_14 = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o {obs_where} AND o.observation_group = 'HIPO'", obs_p)['c']
        near_14 = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o {obs_where} AND o.observation_group = 'Near Miss'", obs_p)['c']
        total_obs_14 = db_fetchone(conn, f"SELECT COUNT(*) as c FROM observations o {obs_where}", obs_p)['c'] or 1
        unsafe_share = sum(unsafe_series) / total_obs_14 * 100 if total_obs_14 else 0

        safety_score = min(30, int(unsafe_share * 0.25) + min(15, int(max(0, unsafe_slope) * 3)) + min(10, hipo_14 * 3) + min(8, near_14 // 2))
        if unsafe_accel > 0.35 and w2_unsafe >= 5:
            safety_score = min(30, safety_score + 8)
        domains['safety'] = {
            'score': safety_score,
            'trend': 'worsening' if unsafe_slope > 0.15 or unsafe_accel > 0.25 else ('improving' if unsafe_slope < -0.1 else 'stable'),
            'summary': f'{sum(unsafe_series)} unsafe/near/HIPO in 14d; HIPO={hipo_14}, near-miss={near_14}; unsafe share {unsafe_share:.0f}%',
        }
        if unsafe_slope > 0.12 and sum(unsafe_series) >= 8:
            days_to_double = None
            if unsafe_slope > 0.01:
                avg_tail = sum(unsafe_series[-5:]) / min(5, len(unsafe_series))
                days_to_double = int(max(3, min(21, (avg_tail / max(0.01, unsafe_slope))))) if avg_tail else None
            predictive_signals.append({
                'type': 'predictive',
                'severity': 'high' if unsafe_slope > 0.25 else 'medium',
                'title': 'Unsafe observation rate trending up',
                'detail': f'14-day slope indicates increasing unsafe/near-miss volume (week-over-week {(unsafe_accel*100):.0f}% change). Early intervention reduces incident probability.',
                'horizon_days': days_to_double or 14,
                'confidence': min(0.88, 0.5 + min(0.35, unsafe_slope * 0.8)),
            })
            preventive_recommendations.append({
                'priority': 1,
                'domain': 'safety',
                'action': 'Trigger focused field leadership walkthrough and toolbox talk on top recurring observation types.',
                'rationale': 'Statistical leading indicator: unsafe observation trajectory positive before incidents.',
            })
        if hipo_14 > 0:
            preventive_recommendations.append({
                'priority': 1,
                'domain': 'safety',
                'action': 'HIPO review board: barrier analysis and verification of controls within 48 hours.',
                'rationale': f'{hipo_14} HIPO-level signal(s) in the analysis window.',
            })

        # --- TWL stress (7d) ---
        twl_where = 'WHERE t.reading_date >= ? AND t.reading_date <= ?'
        twl_p = [d7, target_date]
        if area_id:
            twl_where += ' AND t.area_id = ?'
            twl_p.append(area_id)
        elif division_id:
            twl_where += ''' AND t.area_id IN (SELECT a.id FROM areas a WHERE a.division_id = ?)'''
            twl_p.append(division_id)

        twl_total = db_fetchone(conn, f'SELECT COUNT(*) as c FROM twl_readings t {twl_where}', twl_p)['c'] or 1
        twl_high_c = db_fetchone(conn, f"SELECT COUNT(*) as c FROM twl_readings t {twl_where} AND t.risk_zone = 'high'", twl_p)['c']
        twl_high_pct = (twl_high_c / twl_total) * 100
        env_score = min(15, int(twl_high_pct * 0.12) + twl_high_c * 2)
        domains['environmental'] = {
            'score': env_score,
            'trend': 'stable',
            'summary': f'{twl_high_c} high-risk TWL readings in last 7 days ({twl_high_pct:.0f}% of readings in scope)',
        }
        if twl_high_pct > 15 and twl_total >= 5:
            predictive_signals.append({
                'type': 'predictive',
                'severity': 'high',
                'title': 'Thermal stress exposure elevated',
                'detail': 'Share of high-risk TWL readings suggests heat strain risk for outdoor/heavy work — expect fatigue-related errors if unmitigated.',
                'horizon_days': 3,
                'confidence': 0.72,
            })
            preventive_recommendations.append({
                'priority': 2,
                'domain': 'environmental',
                'action': 'Enforce work-rest cycles, hydration points, and reschedule heavy work outside peak heat.',
                'rationale': 'Environmental leading indicator correlates with safety and health incidents in hot seasons.',
            })

        # --- Cross: attendance + thermal ---
        am_today = db_fetchone(conn, f'''SELECT COUNT(DISTINCT att.employee_no) as c
            FROM attendance att JOIN employees e ON e.employee_no = att.employee_no AND e.project_id = att.project_id
            LEFT JOIN projects p ON p.id = att.project_id LEFT JOIN areas ar ON ar.id = p.area_id
            WHERE att.scan_date = ? AND att.session = 'AM' AND e.active = 1''' + (
            ' AND att.project_id = ?' if project_id else (' AND p.area_id = ?' if area_id else (' AND ar.division_id = ?' if division_id else ''))),
            [target_date] + ([project_id] if project_id else [area_id] if area_id else [division_id] if division_id else []))['c']
        cross_score = 0
        if twl_high_c > 0 and total_emp > 20 and am_today > total_emp * 0.65:
            cross_score = min(10, 6 + twl_high_c)
            predictive_signals.append({
                'type': 'predictive',
                'severity': 'critical',
                'title': 'Compound risk: heat + high site occupancy',
                'detail': f'{am_today} workers on site while TWL readings include high-risk zones — elevated heat illness and human error risk.',
                'horizon_days': 1,
                'confidence': 0.78,
            })
            preventive_recommendations.append({
                'priority': 1,
                'domain': 'cross_cutting',
                'action': 'Consider temporary reduction of non-essential outdoor work; brief supervisors on heat emergency response.',
                'rationale': 'Cross-domain correlation: environmental stress × headcount.',
            })
        domains['cross_cutting'] = {'score': cross_score, 'trend': 'stable', 'summary': 'Correlated attendance × environment signals' if cross_score else 'No compound signal'}

        risk_index = min(100, att_score + health_score + safety_score + env_score + cross_score)
        risk_level = _risk_level_from_index(risk_index)

        preventive_recommendations.sort(key=lambda x: x['priority'])

        severity_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        predictive_signals.sort(key=lambda s: severity_order.get(s.get('severity', 'low'), 9))

    finally:
        conn.close()

    return jsonify({
        'date': target_date,
        'risk_index': risk_index,
        'risk_level': risk_level,
        'domains': domains,
        'predictive_signals': predictive_signals,
        'preventive_recommendations': preventive_recommendations,
        'methodology': 'Rule-based leading indicators + 14d linear trends on attendance and unsafe observations; not a certified ML model.',
    })


# ============================================================
# O&I Observations
# ============================================================

OI_OBSERVATION_TYPES = [
    'Hot Work', 'Safety Devices or Guards', 'Procedures',
    'Personal Protective Equipment', 'Tools Equipment', 'Lifting',
    'Health, Hygiene, Food or Water', 'Safety Signage and Demarcation',
    'Emergency Response', 'Other', 'Housekeeping', 'Driving/Vehicles',
    'Excavation', 'Line of Fire or Pinch Points', 'Working at Height',
    'Workplace Environment', 'Work Planning & Authorisation', 'Supervision',
    'Situational Awareness', 'Toxic/Flammable Gas', 'Confined Space',
    'Isolation/Lockout', 'Improvement Opportunity', 'Manual/Mechanical Handling',
    'Security',
]

OI_OBSERVATION_GROUPS = ['Safe Act', 'Safe Condition', 'Unsafe Act', 'Unsafe Condition', 'Near Miss', 'HIPO']
OI_SEVERITY_UNSAFE = ['Low', 'Medium', 'High', 'Critical']
OI_SEVERITY_SAFE = ['N/A']
OI_RISK_UNSAFE = ['Low', 'Medium', 'High']
OI_RISK_SAFE = ['N/A']
OI_EMPLOYEE_TYPES = ['AON Direct Hire', 'PMC', 'Contractor']


@app.route('/api/observations/meta')
@require_auth
def api_observations_meta():
    return jsonify({
        'observation_types': OI_OBSERVATION_TYPES,
        'observation_groups': OI_OBSERVATION_GROUPS,
        'severity_unsafe': OI_SEVERITY_UNSAFE,
        'severity_safe': OI_SEVERITY_SAFE,
        'risk_unsafe': OI_RISK_UNSAFE,
        'risk_safe': OI_RISK_SAFE,
        'employee_types': OI_EMPLOYEE_TYPES,
    })


@app.route('/api/observations', methods=['POST'])
@require_auth
def api_observation_create():
    data = request.json or {}
    obs_date = data.get('observation_date', date.today().isoformat())
    obs_group = data.get('observation_group', '').strip()
    if not obs_group:
        return jsonify({'success': False, 'message': 'Observation group required'}), 400
    conn = get_db()
    try:
        db_execute(conn, '''INSERT INTO observations
            (division_id, area_id, project_id, observation_date, observer_name,
             observer_designation, observer_discipline, observer_company, employee_type,
             observation_group, observation_type, potential_severity, risk_rating,
             observation_text, corrective_action, intervention, outcome, remarks, recorded_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (data.get('division_id'), data.get('area_id'), data.get('project_id'),
             obs_date, data.get('observer_name', ''), data.get('observer_designation', ''),
             data.get('observer_discipline', ''), data.get('observer_company', ''),
             data.get('employee_type', ''), obs_group, data.get('observation_type', ''),
             data.get('potential_severity', ''), data.get('risk_rating', ''),
             data.get('observation_text', ''), data.get('corrective_action', ''),
             data.get('intervention', ''), data.get('outcome', ''),
             data.get('remarks', ''), request.user['id']))
        conn.commit()
    finally:
        conn.close()
    audit(request.user['id'], 'observation_recorded', f'{obs_group}: {data.get("observation_type", "")}')
    return jsonify({'success': True})


@app.route('/api/observations')
@require_auth
def api_observations_list():
    obs_date = request.args.get('date', '')
    days = request.args.get('days', type=int)
    division_id = request.args.get('division_id', type=int)
    area_id = request.args.get('area_id', type=int)
    project_id = request.args.get('project_id', type=int)

    conn = get_db()
    try:
        query = '''SELECT o.*, d.name as division_name, a.name as area_name, p.name as project_name
            FROM observations o
            LEFT JOIN divisions d ON d.id = o.division_id
            LEFT JOIN areas a ON a.id = o.area_id
            LEFT JOIN projects p ON p.id = o.project_id
            WHERE 1=1'''
        params = []
        if obs_date:
            query += ' AND o.observation_date = ?'
            params.append(obs_date)
        elif days:
            start = (date.today() - timedelta(days=days)).isoformat()
            query += ' AND o.observation_date >= ?'
            params.append(start)
        if division_id:
            query += ' AND o.division_id = ?'
            params.append(division_id)
        if area_id:
            query += ' AND o.area_id = ?'
            params.append(area_id)
        if project_id:
            query += ' AND o.project_id = ?'
            params.append(project_id)
        query += ' ORDER BY o.observation_date DESC, o.created_at DESC LIMIT 500'
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/observations/summary')
@require_auth
def api_observations_summary():
    """Summary for dashboard: totals by group, type, severity for a date range."""
    obs_date = request.args.get('date', '')
    days = request.args.get('days', 30, type=int)
    division_id = request.args.get('division_id', type=int)
    area_id = request.args.get('area_id', type=int)
    project_id = request.args.get('project_id', type=int)

    conn = get_db()
    try:
        where = 'WHERE 1=1'
        params = []
        if obs_date:
            where += ' AND o.observation_date = ?'
            params.append(obs_date)
        else:
            start = (date.today() - timedelta(days=days)).isoformat()
            where += ' AND o.observation_date >= ?'
            params.append(start)
        if division_id:
            where += ' AND o.division_id = ?'
            params.append(division_id)
        if area_id:
            where += ' AND o.area_id = ?'
            params.append(area_id)
        if project_id:
            where += ' AND o.project_id = ?'
            params.append(project_id)

        base = f'FROM observations o {where}'
        total = db_fetchone(conn, f'SELECT COUNT(*) as c {base}', params)['c']
        by_group = db_fetchall(conn, f'SELECT observation_group, COUNT(*) as c {base} GROUP BY observation_group ORDER BY c DESC', params)
        by_type = db_fetchall(conn, f"SELECT observation_type, COUNT(*) as c {base} AND observation_type IS NOT NULL AND observation_type != '' GROUP BY observation_type ORDER BY c DESC LIMIT 10", params)
        by_severity = db_fetchall(conn, f"SELECT potential_severity, COUNT(*) as c {base} AND potential_severity IS NOT NULL AND potential_severity != '' GROUP BY potential_severity ORDER BY c DESC", params)
        by_date = db_fetchall(conn, f'SELECT observation_date, COUNT(*) as c {base} GROUP BY observation_date ORDER BY observation_date', params)
        safe_count = db_fetchone(conn, f"SELECT COUNT(*) as c {base} AND observation_group IN ('Safe Act','Safe Condition')", params)['c']
        unsafe_count = total - safe_count
    finally:
        conn.close()
    return jsonify({
        'total': total,
        'safe': safe_count,
        'unsafe': unsafe_count,
        'by_group': [dict(r) for r in by_group],
        'by_type': [dict(r) for r in by_type],
        'by_severity': [dict(r) for r in by_severity],
        'by_date': [dict(r) for r in by_date],
    })


@app.route('/api/observations/insights')
@require_auth
def api_observations_insights():
    """Detect O&I anomalies: rising trends, discipline hotspots, recurring types, HIPO clusters."""
    days = request.args.get('days', 60, type=int)
    division_id = request.args.get('division_id', type=int)
    area_id = request.args.get('area_id', type=int)
    project_id = request.args.get('project_id', type=int)

    conn = get_db()
    try:
        where = 'WHERE 1=1'
        params = []
        start = (date.today() - timedelta(days=days)).isoformat()
        where += ' AND o.observation_date >= ?'
        params.append(start)
        if division_id:
            where += ' AND o.division_id = ?'
            params.append(division_id)
        if area_id:
            where += ' AND o.area_id = ?'
            params.append(area_id)
        if project_id:
            where += ' AND o.project_id = ?'
            params.append(project_id)

        base = f'FROM observations o {where}'
        insights = []

        # 1. Recurring observation types (top repeated)
        recurring = db_fetchall(conn, f"""
            SELECT observation_type, COUNT(*) as c,
                   observation_group
            {base} AND observation_type IS NOT NULL AND observation_type != ''
            GROUP BY observation_type
            HAVING COUNT(*) >= 3
            ORDER BY c DESC LIMIT 10
        """, params)

        # 2. Weekly trend comparison for unsafe observations
        mid = (date.today() - timedelta(days=days//2)).isoformat()
        first_half = db_fetchone(conn, f"""
            SELECT COUNT(*) as c {base}
            AND observation_group NOT IN ('Safe Act','Safe Condition')
            AND o.observation_date < ?
        """, params + [mid])['c']
        second_half = db_fetchone(conn, f"""
            SELECT COUNT(*) as c {base}
            AND observation_group NOT IN ('Safe Act','Safe Condition')
            AND o.observation_date >= ?
        """, params + [mid])['c']

        if first_half > 0 and second_half > first_half * 1.2:
            pct_rise = round(((second_half - first_half) / first_half) * 100)
            insights.append({
                'severity': 'high',
                'category': 'Trend',
                'title': f'Unsafe observations rising ({pct_rise}% increase)',
                'detail': f'Recent {days//2} days: {second_half} unsafe vs prior {days//2} days: {first_half}',
                'metric': f'+{pct_rise}%'
            })

        # 3. Discipline hotspots
        by_disc = db_fetchall(conn, f"""
            SELECT observer_discipline, COUNT(*) as c,
                   SUM(CASE WHEN observation_group NOT IN ('Safe Act','Safe Condition') THEN 1 ELSE 0 END) as unsafe_c
            {base} AND observer_discipline IS NOT NULL AND observer_discipline != ''
            GROUP BY observer_discipline
            ORDER BY unsafe_c DESC
        """, params)

        total_obs = db_fetchone(conn, f'SELECT COUNT(*) as c {base}', params)['c'] or 1
        for disc in by_disc[:5]:
            d_name = disc['observer_discipline']
            d_total = disc['c']
            d_unsafe = disc['unsafe_c']
            share = round((d_total / total_obs) * 100)
            unsafe_rate = round((d_unsafe / d_total) * 100) if d_total > 0 else 0
            if d_unsafe >= 3 and unsafe_rate > 30:
                insights.append({
                    'severity': 'high' if unsafe_rate > 50 else 'medium',
                    'category': 'Discipline',
                    'title': f'{d_name}: {unsafe_rate}% unsafe rate',
                    'detail': f'{d_unsafe}/{d_total} observations are unsafe/near-miss/HIPO ({share}% of all observations)',
                    'metric': f'{unsafe_rate}%'
                })

        # 4. HIPO events
        hipos = db_fetchall(conn, f"""
            SELECT o.observation_date, o.observation_type, o.observer_discipline,
                   d.name as division_name
            {base}
            LEFT JOIN divisions d ON d.id = o.division_id
            AND o.observation_group = 'HIPO'
            ORDER BY o.observation_date DESC
        """, params)
        if hipos:
            insights.append({
                'severity': 'critical',
                'category': 'HIPO',
                'title': f'{len(hipos)} High Potential (HIPO) events detected',
                'detail': f'Most recent: {hipos[0].get("observation_type","N/A")} on {hipos[0].get("observation_date","")}',
                'metric': str(len(hipos))
            })

        # 5. Type-specific rising trends
        for rec in recurring[:5]:
            obs_type = rec['observation_type']
            tc = rec['c']
            first_t = db_fetchone(conn, f"""
                SELECT COUNT(*) as c {base}
                AND observation_type = ? AND o.observation_date < ?
            """, params + [obs_type, mid])['c']
            second_t = db_fetchone(conn, f"""
                SELECT COUNT(*) as c {base}
                AND observation_type = ? AND o.observation_date >= ?
            """, params + [obs_type, mid])['c']
            if first_t > 0 and second_t > first_t * 1.3:
                pct = round(((second_t - first_t) / first_t) * 100)
                insights.append({
                    'severity': 'medium',
                    'category': 'Recurring',
                    'title': f'"{obs_type}" observations increasing (+{pct}%)',
                    'detail': f'Total: {tc} over {days} days. Recent half: {second_t} vs prior: {first_t}',
                    'metric': str(tc)
                })

        # 6. Area concentration
        by_area = db_fetchall(conn, f"""
            SELECT a.name as area_name, COUNT(*) as c,
                   SUM(CASE WHEN o.observation_group NOT IN ('Safe Act','Safe Condition') THEN 1 ELSE 0 END) as unsafe_c
            {base}
            LEFT JOIN areas a ON a.id = o.area_id
            AND a.name IS NOT NULL
            GROUP BY a.name
            ORDER BY unsafe_c DESC
        """, params)
        for a in by_area[:3]:
            a_name = a.get('area_name', 'Unknown')
            if not a_name:
                continue
            a_unsafe = a['unsafe_c']
            a_total = a['c']
            if a_unsafe >= 5:
                insights.append({
                    'severity': 'medium',
                    'category': 'Area',
                    'title': f'{a_name}: {a_unsafe} unsafe observations',
                    'detail': f'{a_total} total observations in this area, {a_unsafe} are unsafe/near-miss/HIPO',
                    'metric': str(a_unsafe)
                })

        # 7. Safe observation ratio
        safe_total = db_fetchone(conn, f"""
            SELECT COUNT(*) as c {base}
            AND observation_group IN ('Safe Act','Safe Condition')
        """, params)['c']
        if total_obs > 10:
            safe_pct = round((safe_total / total_obs) * 100)
            if safe_pct < 50:
                insights.append({
                    'severity': 'medium',
                    'category': 'Ratio',
                    'title': f'Low safe observation ratio: {safe_pct}%',
                    'detail': f'Only {safe_total} of {total_obs} observations are safe acts/conditions. Target: >60%',
                    'metric': f'{safe_pct}%'
                })

        # Sort by severity
        sev_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        insights.sort(key=lambda a: sev_order.get(a['severity'], 9))

    finally:
        conn.close()

    return jsonify({
        'count': len(insights),
        'insights': insights,
        'recurring': [dict(r) for r in recurring] if recurring else [],
        'by_discipline': [dict(r) for r in by_disc] if by_disc else [],
    })


# ============================================================
# QR / Export / Audit
# ============================================================

@app.route('/api/qrcodes/batch')
@require_auth
def api_qrcodes_batch():
    conn = get_db()
    try:
        project_id = request.args.get('project_id', '')
        area_id = request.args.get('area_id', type=int)
        division_id = request.args.get('division_id', type=int)
        query = '''SELECT e.project_id, e.employee_no, e.name, e.designation, e.discipline, e.work_location, p.name as project_name
            FROM employees e
            JOIN projects p ON p.id = e.project_id
            LEFT JOIN areas a ON a.id = p.area_id
            WHERE e.active = 1'''
        params = []
        if project_id:
            query += ' AND e.project_id = ?'
            params.append(project_id)
        if area_id:
            query += ' AND p.area_id = ?'
            params.append(area_id)
        if division_id:
            query += ' AND a.division_id = ?'
            params.append(division_id)
        query, params = apply_project_filter(conn, query, params, request.user, 'e')
        query += ' ORDER BY e.name'
        employees = db_fetchall(conn, query, params)
    finally:
        conn.close()

    # Unique QR per person: encode project_id|employee_no so each code is globally unique
    results = []
    for emp in employees:
        qr_payload = f"{emp['project_id']}|{emp['employee_no']}"
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=6, border=3)
        qr.add_data(qr_payload)
        qr.make(fit=True)
        buf = BytesIO()
        qr.make_image(fill_color="black", back_color="white").save(buf, format='PNG')
        results.append({**emp, 'qr_base64': base64.b64encode(buf.getvalue()).decode()})
    return jsonify(results)


@app.route('/api/qrcodes/single')
@require_auth
def api_qrcode_single():
    """Download a single QR code PNG for one employee."""
    employee_no = request.args.get('employee_no', '').strip()
    project_id = request.args.get('project_id', '').strip()
    if not employee_no or not project_id:
        return jsonify({'error': 'employee_no and project_id required'}), 400
    qr_payload = f"{project_id}|{employee_no}"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
    qr.add_data(qr_payload)
    qr.make(fit=True)
    buf = BytesIO()
    qr.make_image(fill_color="black", back_color="white").save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png', as_attachment=True, download_name=f'QR_{employee_no}.png')


@app.route('/api/export/download-token', methods=['POST'])
@require_auth
def api_download_token():
    conn = get_db()
    try:
        tok = secrets.token_urlsafe(32)
        db_execute(conn, "INSERT INTO download_tokens (user_id, token, expires_at) VALUES (?, ?, datetime('now', '+5 minutes'))",
                   (request.user['id'], tok))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'token': tok})


@app.route('/api/export/attendance')
def api_export():
    dl_token = request.args.get('dl_token', '')
    if not dl_token:
        return jsonify({'error': 'Missing download token'}), 401

    conn = get_db()
    try:
        if USE_POSTGRES:
            tok_row = db_fetchone(conn, '''
                SELECT d.user_id, u.* FROM download_tokens d JOIN users u ON u.id = d.user_id
                WHERE d.token = ? AND d.expires_at > NOW()
            ''', (dl_token,))
        else:
            tok_row = db_fetchone(conn, '''
                SELECT d.user_id, u.* FROM download_tokens d JOIN users u ON u.id = d.user_id
                WHERE d.token = ? AND d.expires_at > datetime('now')
            ''', (dl_token,))

        if not tok_row:
            return jsonify({'error': 'Invalid or expired token'}), 401

        db_execute(conn, 'DELETE FROM download_tokens WHERE token = ?', (dl_token,))
        conn.commit()

        scan_date = request.args.get('date', date.today().isoformat())
        project_id = request.args.get('project_id')

        query = '''SELECT e.employee_no, e.name, e.designation, e.discipline, e.nationality, e.work_location,
            e.subcontractor, p.name as project_name, a.session, a.scan_date, a.scanned_at, a.supervisor_name, a.scanner_email, a.scanner_designation
            FROM attendance a JOIN employees e ON e.id = a.employee_id JOIN projects p ON p.id = a.project_id WHERE a.scan_date = ?'''
        params = [scan_date]
        if project_id:
            query += ' AND a.project_id = ?'
            params.append(project_id)
        query, params = apply_project_filter(conn, query, params, tok_row)
        query += ' ORDER BY p.name, e.name'
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Project', 'Employee No', 'Name', 'Designation', 'Discipline', 'Nationality',
                     'Work Location', 'Sub-contractor', 'Session', 'Date', 'Scanned At', 'Supervisor', 'Scanner Email', 'Scanner Designation'])
    for r in rows:
        writer.writerow([r['project_name'], r['employee_no'], r['name'], r['designation'], r['discipline'],
                        r['nationality'], r['work_location'], r['subcontractor'], r['session'], r['scan_date'],
                        r['scanned_at'], r['supervisor_name'], r.get('scanner_email') or '', r.get('scanner_designation') or ''])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name=f'attendance_{scan_date}.csv')


# Full roster export: all required columns (Agreement No., Name, Nationality, DOB, Designation, Physical Work Location, Residing Camp, Employee No., Qualification, Date of Joining, Date of Deployment, Latest Periodic Medical, Discipline, Sub-contractor, Remarks)
@app.route('/api/export/roster')
def api_export_roster():
    dl_token = request.args.get('dl_token', '')
    if not dl_token:
        return jsonify({'error': 'Missing download token'}), 401
    conn = get_db()
    try:
        if USE_POSTGRES:
            tok_row = db_fetchone(conn, '''
                SELECT d.user_id, u.* FROM download_tokens d JOIN users u ON u.id = d.user_id
                WHERE d.token = ? AND d.expires_at > NOW()
            ''', (dl_token,))
        else:
            tok_row = db_fetchone(conn, '''
                SELECT d.user_id, u.* FROM download_tokens d JOIN users u ON u.id = d.user_id
                WHERE d.token = ? AND d.expires_at > datetime('now')
            ''', (dl_token,))
        if not tok_row:
            return jsonify({'error': 'Invalid or expired token'}), 401
        db_execute(conn, 'DELETE FROM download_tokens WHERE token = ?', (dl_token,))
        conn.commit()
        project_id = request.args.get('project_id', '')
        query = '''SELECT e.srl, e.agreement_no, e.asset_name, e.contractor, e.name, e.nationality, e.dob, e.designation,
            e.age, e.eid_passport, e.fieldglass_status,
            e.work_location, e.camp_name, e.employee_no, e.qualification, e.date_joining,
            e.date_deployment, e.medical_date, e.discipline, e.subcontractor,
            e.medical_frequency, e.last_medical_date, e.next_medical_due, e.medical_result,
            e.chronic_condition, e.chronic_treated, e.general_feeling, e.remarks,
            p.name as project_name
            FROM employees e JOIN projects p ON p.id = e.project_id WHERE e.active = 1'''
        params = []
        if project_id:
            query += ' AND e.project_id = ?'
            params.append(project_id)
        query, params = apply_project_filter(conn, query, params, tok_row)
        query += ' ORDER BY p.name, e.name'
        rows = db_fetchall(conn, query, params)
    finally:
        conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'SL#', 'Agreement No.', 'Asset Name', 'Contractor', 'Project Name',
        'Employee Full Name', 'Designation', 'Nationality', 'DOB', 'Age',
        'EID / Passport #', 'Fieldglass Status', 'Physical Work Location',
        'Residing Camp Name & Location', 'Employee No. Contractor Ref.',
        'Qualification', 'Date of Joining', 'Date of Deployment',
        'Latest Medical Date', 'Discipline', 'Sub-contractor Manpower Supplier',
        'Frequency of Medical', 'Date of Last Medical', 'Date of Next Medical Due',
        'Result (Fit/Unfit)', 'Chronic Condition', 'Chronic Treated?',
        'How Do You Generally Feel?', 'Remarks'
    ])
    for r in rows:
        writer.writerow([
            r.get('srl') or '', r.get('agreement_no') or '', r.get('asset_name') or '',
            r.get('contractor') or '', r.get('project_name') or '', r.get('name') or '',
            r.get('designation') or '', r.get('nationality') or '', r.get('dob') or '',
            r.get('age') or '', r.get('eid_passport') or '', r.get('fieldglass_status') or '',
            r.get('work_location') or '', r.get('camp_name') or '', r.get('employee_no') or '',
            r.get('qualification') or '', r.get('date_joining') or '', r.get('date_deployment') or '',
            r.get('medical_date') or '', r.get('discipline') or '', r.get('subcontractor') or '',
            r.get('medical_frequency') or '', r.get('last_medical_date') or '',
            r.get('next_medical_due') or '', r.get('medical_result') or '',
            r.get('chronic_condition') or '', r.get('chronic_treated') or '',
            r.get('general_feeling') or '', r.get('remarks') or ''
        ])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='roster_full.csv')


@app.route('/api/scanner-status')
@require_role('executive', 'admin')
def api_scanner_status():
    """Return all scanner users with scan activity, approval status, and security info."""
    conn = get_db()
    try:
        scanners = db_fetchall(conn, '''
            SELECT u.id, u.username, u.display_name, u.email, u.designation,
                   u.active, u.created_at, u.last_login, u.failed_attempts, u.locked_until
            FROM users u WHERE u.role = 'scanner' ORDER BY u.display_name
        ''')
        result = []
        today = date.today().isoformat()
        for s in scanners:
            uid = s['id']
            projects = db_fetchall(conn, '''
                SELECT p.id, p.name, a2.name as area_name FROM user_project_access upa
                JOIN projects p ON p.id = upa.project_id
                LEFT JOIN areas a2 ON a2.id = p.area_id
                WHERE upa.user_id = ?
            ''', (uid,))

            last_scan = db_fetchone(conn, '''
                SELECT scan_date, session, scanned_at FROM attendance
                WHERE supervisor_id = ? ORDER BY scanned_at DESC LIMIT 1
            ''', (uid,))

            today_count = db_fetchone(conn, '''
                SELECT COUNT(*) as c FROM attendance
                WHERE supervisor_id = ? AND scan_date = ?
            ''', (uid, today))

            total_scans = db_fetchone(conn, '''
                SELECT COUNT(*) as c FROM attendance WHERE supervisor_id = ?
            ''', (uid,))

            days_active = db_fetchone(conn, '''
                SELECT COUNT(DISTINCT scan_date) as c FROM attendance WHERE supervisor_id = ?
            ''', (uid,))

            row = dict(s)
            row['projects'] = [dict(p) for p in projects]
            row['last_scan'] = dict(last_scan) if last_scan else None
            row['today_scans'] = today_count['c'] if today_count else 0
            row['total_scans'] = total_scans['c'] if total_scans else 0
            row['days_active'] = days_active['c'] if days_active else 0
            result.append(row)
    finally:
        conn.close()
    return jsonify(result)


@app.route('/api/audit')
@require_role('executive', 'admin')
def api_audit():
    conn = get_db()
    try:
        rows = db_fetchall(conn, '''SELECT a.*, u.display_name as user_name FROM audit_log a
            LEFT JOIN users u ON u.id = a.user_id ORDER BY a.created_at DESC LIMIT 200''')
    finally:
        conn.close()
    return jsonify(rows)


# ============================================================
# ROSTER IMPORT
# ============================================================

def import_roster(csv_path=None, project_name=None):
    if csv_path is None:
        csv_path = ROSTER_CSV
    if not os.path.exists(csv_path):
        return 0, "CSV file not found"

    if not project_name:
        project_name = os.path.splitext(os.path.basename(csv_path))[0].replace('_', ' ').title()

    conn = get_db()
    try:
        existing = db_fetchone(conn, 'SELECT id FROM projects WHERE name = ?', (project_name,))
        if not existing:
            db_execute(conn, 'INSERT INTO projects (name) VALUES (?)', (project_name,))
            conn.commit()
        project = db_fetchone(conn, 'SELECT id FROM projects WHERE name = ?', (project_name,))
        project_id = project['id']

        count = 0
        sites_found = set()

        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            content = f.read()
        content = content.replace('\n(Site / City Name) ', '').replace('\nLocation\n', ' Location,').replace('\nContractor Ref.', ' Contractor Ref.')

        lines = content.strip().split('\n')
        fixed = [lines[0]] + [l for l in lines[1:] if l.strip()]
        reader = csv.reader(io.StringIO('\n'.join(fixed)))
        next(reader)

        agr = None
        for row in reader:
            if len(row) < 15 or not row[0].strip():
                continue
            try:
                vals = [c.strip() for c in row[:16]] + ([''] if len(row) <= 15 else [row[15].strip()])
                srl, agreement_no, name, nationality, dob, designation = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5]
                work_location, camp_name, employee_no = vals[6], vals[7], vals[8]
                qualification, date_joining, date_deployment = vals[9], vals[10], vals[11]
                medical_date, discipline, subcontractor = vals[12], vals[13], vals[14]
                remarks = vals[15] if len(vals) > 15 else ''

                if not employee_no or not name:
                    continue
                if not agr and agreement_no:
                    agr = agreement_no
                    db_execute(conn, 'UPDATE projects SET agreement_no = ? WHERE id = ?', (agreement_no, project_id))

                sites_found.add(work_location)

                existing_emp = db_fetchone(conn, 'SELECT id FROM employees WHERE employee_no = ? AND project_id = ?', (employee_no, project_id))
                if existing_emp:
                    db_execute(conn, '''UPDATE employees SET srl=?, agreement_no=?, name=?, nationality=?, dob=?,
                        designation=?, work_location=?, camp_name=?, qualification=?, date_joining=?,
                        date_deployment=?, medical_date=?, discipline=?, subcontractor=?, remarks=?
                        WHERE employee_no = ? AND project_id = ?''',
                        (srl, agreement_no, name, nationality, dob, designation, work_location, camp_name,
                         qualification, date_joining, date_deployment, medical_date, discipline, subcontractor,
                         remarks, employee_no, project_id))
                else:
                    db_execute(conn, '''INSERT INTO employees (srl, agreement_no, name, nationality, dob, designation,
                        work_location, camp_name, employee_no, qualification, date_joining, date_deployment,
                        medical_date, discipline, subcontractor, remarks, project_id)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                        (srl, agreement_no, name, nationality, dob, designation, work_location, camp_name,
                         employee_no, qualification, date_joining, date_deployment, medical_date, discipline,
                         subcontractor, remarks, project_id))
                count += 1
            except Exception as e:
                print(f"Import error: {e}")

        for sn in sites_found:
            if sn:
                existing_site = db_fetchone(conn, 'SELECT id FROM sites WHERE name = ? AND project_id = ?', (sn, project_id))
                if not existing_site:
                    db_execute(conn, 'INSERT INTO sites (name, project_id) VALUES (?, ?)', (sn, project_id))

        conn.commit()
    finally:
        conn.close()
    return count, f"Imported {count} employees into '{project_name}'"


def _col_index(header_row, *names):
    """Find column index by header name (case-insensitive, partial match)."""
    for i, cell in enumerate(header_row):
        raw = getattr(cell, 'value', cell) or ''
        val = str(raw).strip().lower().replace('\n', ' ')
        for n in names:
            if n.lower() in val or val in n.lower():
                return i
    return -1


SHEET_TO_PROJECT = {}

def _cell_str(row, idx):
    if idx < 0 or idx >= len(row) or row[idx] is None:
        return ''
    v = row[idx]
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def import_excel(xlsx_path, area_id=None):
    """Import sheets from 16-column or 29-column Excel template. Auto-detects format. Projects created under given area_id."""
    try:
        import openpyxl
    except ImportError:
        return 0, "Install openpyxl: pip install openpyxl"

    if not os.path.exists(xlsx_path):
        return 0, "Excel file not found"
    if not area_id:
        return 0, "Area is required for import"

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    conn = get_db()
    total = 0
    messages = []
    # Resolve area name for project naming
    area_row = db_fetchone(conn, 'SELECT name FROM areas WHERE id = ?', (area_id,))
    area_label = (area_row['name'] if isinstance(area_row, dict) else area_row[0]) if area_row else ''
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Read project name from cell C3; fall back to sheet name
            c3_val = None
            try:
                c3_val = ws['C3'].value
            except Exception:
                pass
            raw = str(c3_val).strip() if c3_val else sheet_name.strip()
            project_name = raw

            # Use provided area_id for new projects
            existing = db_fetchone(conn, 'SELECT id FROM projects WHERE area_id = ? AND name = ?', (area_id, project_name))
            if not existing:
                db_execute(conn, 'INSERT INTO projects (area_id, name, active) VALUES (?, ?, 1)', (area_id, project_name))
                conn.commit()
            proj = db_fetchone(conn, 'SELECT id FROM projects WHERE area_id = ? AND name = ?', (area_id, project_name))
            project_id = proj['id']
            site_name = project_name
            existing_site = db_fetchone(conn, 'SELECT id FROM sites WHERE name = ? AND project_id = ?', (site_name, project_id))
            if not existing_site:
                db_execute(conn, 'INSERT INTO sites (name, project_id, active) VALUES (?, ?, 1)', (site_name, project_id))
                conn.commit()

            # Detect format: 29-col welfare template vs 16-col standard
            # 29-col header row has "ASSET NAME" or "CONTRACTOR" or "EID"
            is_29col = False
            header_row_num = 1
            for test_row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
                header_text = ' '.join(str(c or '').lower() for c in test_row)
                if 'asset name' in header_text or 'fieldglass' in header_text or 'eid' in header_text:
                    is_29col = True
                    break
                header_row_num += 1
            data_start = header_row_num + 1 if is_29col else 2

            count = 0
            sites_found = set()
            for row in ws.iter_rows(min_row=data_start, values_only=True):
                if not row or len(row) < 9:
                    continue
                try:
                    if is_29col:
                        # 29-col: 0:SL#, 1:Agreement, 2:Asset, 3:Contractor, 4:Project, 5:Name, 6:Designation,
                        # 7:Nationality, 8:DOB, 9:Age, 10:EID, 11:Fieldglass, 12:Work Location,
                        # 13:Camp, 14:Employee No, 15:Qualification, 16:Date Joining, 17:Date Deployment,
                        # 18:Medical Date, 19:Discipline, 20:Subcontractor, 21:Med Frequency,
                        # 22:Last Medical, 23:Next Medical, 24:Result, 25:Chronic, 26:Chronic Treated,
                        # 27:General Feeling, 28:Remarks
                        name = _cell_str(row, 5)
                        employee_no = _cell_str(row, 14)
                        if not name or not employee_no:
                            continue
                        skip = employee_no.lower()
                        if skip in ('employee no.', 'contractor ref.', 'employee no', 'none', ''):
                            continue
                        srl = _cell_str(row, 0)
                        agreement_no = _cell_str(row, 1)
                        asset_name = _cell_str(row, 2)
                        contractor_col = _cell_str(row, 3)
                        designation = _cell_str(row, 6)
                        nationality = _cell_str(row, 7)
                        dob = _cell_str(row, 8)
                        age = _cell_str(row, 9)
                        eid_passport = _cell_str(row, 10)
                        fieldglass_status = _cell_str(row, 11)
                        work_location = _cell_str(row, 12) or site_name
                        camp_name = _cell_str(row, 13)
                        qualification = _cell_str(row, 15)
                        date_joining = _cell_str(row, 16)
                        date_deployment = _cell_str(row, 17)
                        medical_date = _cell_str(row, 18)
                        discipline = _cell_str(row, 19)
                        subcontractor = _cell_str(row, 20)
                        medical_frequency = _cell_str(row, 21)
                        last_medical_date = _cell_str(row, 22)
                        next_medical_due = _cell_str(row, 23)
                        medical_result = _cell_str(row, 24)
                        chronic_condition = _cell_str(row, 25)
                        chronic_treated = _cell_str(row, 26)
                        general_feeling = _cell_str(row, 27)
                        remarks = _cell_str(row, 28)
                    else:
                        # 16-col standard
                        name = _cell_str(row, 2)
                        employee_no = _cell_str(row, 8)
                        if not name or not employee_no:
                            continue
                        skip = employee_no.lower()
                        if skip in ('employee no.', 'contractor ref.', 'employee no', 'none', ''):
                            continue
                        srl = _cell_str(row, 0)
                        agreement_no = _cell_str(row, 1)
                        nationality = _cell_str(row, 3)
                        dob = _cell_str(row, 4)
                        designation = _cell_str(row, 5)
                        work_location = _cell_str(row, 6) or site_name
                        camp_name = _cell_str(row, 7)
                        qualification = _cell_str(row, 9)
                        date_joining = _cell_str(row, 10)
                        date_deployment = _cell_str(row, 11)
                        medical_date = _cell_str(row, 12)
                        discipline = _cell_str(row, 13)
                        subcontractor = _cell_str(row, 14)
                        remarks = _cell_str(row, 15)
                        asset_name = contractor_col = age = eid_passport = fieldglass_status = ''
                        medical_frequency = last_medical_date = next_medical_due = medical_result = ''
                        chronic_condition = chronic_treated = general_feeling = ''

                    existing_emp = db_fetchone(conn, 'SELECT id FROM employees WHERE employee_no = ? AND project_id = ?', (employee_no, project_id))
                    if existing_emp:
                        db_execute(conn, '''UPDATE employees SET srl=?, agreement_no=?, name=?, nationality=?, dob=?,
                            designation=?, work_location=?, camp_name=?, qualification=?, date_joining=?,
                            date_deployment=?, medical_date=?, discipline=?, subcontractor=?, remarks=?,
                            asset_name=?, contractor=?, age=?, eid_passport=?, fieldglass_status=?,
                            medical_frequency=?, last_medical_date=?, next_medical_due=?, medical_result=?,
                            chronic_condition=?, chronic_treated=?, general_feeling=?
                            WHERE employee_no = ? AND project_id = ?''',
                            (srl, agreement_no, name, nationality, dob, designation, work_location, camp_name,
                             qualification, date_joining, date_deployment, medical_date, discipline, subcontractor,
                             remarks, asset_name, contractor_col, age, eid_passport, fieldglass_status,
                             medical_frequency, last_medical_date, next_medical_due, medical_result,
                             chronic_condition, chronic_treated, general_feeling,
                             employee_no, project_id))
                    else:
                        db_execute(conn, '''INSERT INTO employees (srl, agreement_no, name, nationality, dob, designation,
                            work_location, camp_name, employee_no, qualification, date_joining, date_deployment,
                            medical_date, discipline, subcontractor, remarks,
                            asset_name, contractor, age, eid_passport, fieldglass_status,
                            medical_frequency, last_medical_date, next_medical_due, medical_result,
                            chronic_condition, chronic_treated, general_feeling, project_id)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                            (srl, agreement_no, name, nationality, dob, designation, work_location, camp_name,
                             employee_no, qualification, date_joining, date_deployment, medical_date, discipline,
                             subcontractor, remarks, asset_name, contractor_col, age, eid_passport, fieldglass_status,
                             medical_frequency, last_medical_date, next_medical_due, medical_result,
                             chronic_condition, chronic_treated, general_feeling, project_id))
                    count += 1
                    if work_location and work_location != site_name:
                        sites_found.add(work_location)
                except Exception as e:
                    print(f"Excel row error ({sheet_name}): {e}")

            for sn in sites_found:
                if sn:
                    ex = db_fetchone(conn, 'SELECT id FROM sites WHERE name = ? AND project_id = ?', (sn, project_id))
                    if not ex:
                        db_execute(conn, 'INSERT INTO sites (name, project_id, active) VALUES (?, ?, 1)', (sn, project_id))
            conn.commit()
            total += count
            messages.append(f"{project_name}: {count} people")
        wb.close()
    finally:
        conn.close()
    return total, ("Excel import: " + "; ".join(messages)) if messages else "No data imported"


# ============================================================
# STARTUP (local dev only)
# ============================================================

if __name__ == '__main__':
    import socket
    import ssl as _ssl

    if os.path.exists(ROSTER_CSV):
        conn = get_db()
        try:
            ec = db_fetchone(conn, 'SELECT COUNT(*) as c FROM employees')['c']
        finally:
            conn.close()
        if ec == 0:
            print("Importing roster...")
            cnt, msg = import_roster()
            print(f"  {msg}")
        else:
            print(f"  {ec} employees in database")

    local_ip = socket.gethostbyname(socket.gethostname())
    cert_dir = os.path.join(os.path.dirname(__file__), 'certs')
    cert_file = os.path.join(cert_dir, 'cert.pem')
    key_file = os.path.join(cert_dir, 'key.pem')
    use_ssl = os.path.exists(cert_file) and os.path.exists(key_file)

    if not use_ssl:
        try:
            os.makedirs(cert_dir, exist_ok=True)
            from cryptography import x509
            from cryptography.x509.oid import NameOID
            from cryptography.hazmat.primitives import hashes, serialization
            from cryptography.hazmat.primitives.asymmetric import rsa
            from cryptography.hazmat.backends import default_backend
            key = rsa.generate_private_key(public_exponent=65537, key_size=2048, backend=default_backend())
            name = x509.Name([x509.NameAttribute(NameOID.COMMON_NAME, 'localhost')])
            cert = (
                x509.CertificateBuilder()
                .subject_name(name)
                .issuer_name(name)
                .public_key(key.public_key())
                .serial_number(x509.random_serial_number())
                .not_valid_before(datetime.now(timezone.utc))
                .not_valid_after(datetime.now(timezone.utc) + timedelta(days=365))
                .add_extension(x509.SubjectAlternativeName([x509.DNSName('localhost'), x509.DNSName('127.0.0.1')]), critical=False)
                .sign(key, hashes.SHA256(), default_backend())
            )
            with open(key_file, 'wb') as f:
                f.write(key.private_bytes(encoding=serialization.Encoding.PEM, format=serialization.PrivateFormat.TraditionalOpenSSL, encryption_algorithm=serialization.NoEncryption()))
            with open(cert_file, 'wb') as f:
                f.write(cert.public_bytes(serialization.Encoding.PEM))
            use_ssl = True
            print("  Generated self-signed HTTPS cert in ./certs/")
        except Exception as e:
            print(f"  HTTPS cert generation skipped: {e}")
            print("  Run with http:// (or add cert.pem + key.pem to ./certs/ for HTTPS)")

    proto = 'https' if use_ssl else 'http'
    print(f"\n{'='*44}")
    print(f"  PTC POB Tracker - Cloud Ready")
    print(f"{'='*44}")
    print(f"  {proto}://localhost:5000")
    print(f"  {proto}://{local_ip}:5000")
    print(f"  DB: {'PostgreSQL' if USE_POSTGRES else 'SQLite (local)'}")
    print(f"{'='*44}\n")

    ssl_ctx = None
    if use_ssl:
        ssl_ctx = _ssl.SSLContext(_ssl.PROTOCOL_TLS_SERVER)
        ssl_ctx.load_cert_chain(cert_file, key_file)

    app.run(host='0.0.0.0', port=5000, debug=True, ssl_context=ssl_ctx)
